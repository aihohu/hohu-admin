"""文件解析器 — spec §16 v1.5+ SR-24

实现 spec §16.3 的 FileParser 协议 + Excel/CSV 两个解析器。PDF / Word 留 v1.6+
（需要 pdfplumber / python-docx 依赖，且业务场景占比 < 10%）。

关键约束（spec §16.1）：
  - 文件 raw bytes 永不进 LLM context（仅返回结构化摘要）
  - 解析器上限：Excel 50MB / CSV 10MB
  - 摘要只含 rows / columns / preview(前 3 行)
  - cell 一律 stringify（防 LLM 看到 datetime / Decimal 等 PydanticAI 不支持的类型）
  - 超限抛 BusinessRuleException(error_code="AI_FILE_TOO_LARGE")
  - 不支持 MIME 抛 BusinessRuleException(error_code="AI_FILE_TYPE_UNSUPPORTED")

同步 IO（openpyxl / csv）用 asyncio.to_thread 包装，避免阻塞事件循环。
"""

import asyncio
import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from app.core.exceptions import BusinessRuleException

# spec §16.1: 摘要只含前 3 行预览（避免大文件 LLM 上下文爆炸）
PREVIEW_ROW_LIMIT = 3


@dataclass(frozen=True)
class FileParseResult:
    """文件解析结构化摘要（传给 LLM 的全部信息）

    raw bytes 不在返回值内（spec §16.1 硬约束）。LLM 收到摘要后，如需用完整数据
    应调具体业务 tool（如 user.batch_create）继续处理，业务 tool 内部自行读文件。
    """

    rows: int
    """总行数（不含表头）"""

    columns: list[str]
    """列名列表（按出现顺序）"""

    preview: list[dict[str, str]]
    """前 PREVIEW_ROW_LIMIT 行预览，key=列名 value=stringify(cell)"""

    parser: str
    """解析器标识，如 'ExcelParser' / 'CsvParser'"""

    file_size: int
    """原始文件大小（字节，审计用）"""


class FileParser(Protocol):
    """spec §16.3 文件解析器协议"""

    mime_types: tuple[str, ...]
    max_bytes: int

    async def parse(self, file_path: Path) -> FileParseResult: ...


# ============ 工具函数 ============


def _row_to_dict(columns: list[str], row: tuple) -> dict[str, str]:
    """行数据按列名拼 dict，cell stringify + None → ''"""
    result: dict[str, str] = {}
    for i, col in enumerate(columns):
        value = row[i] if i < len(row) else None
        result[col] = "" if value is None else str(value)
    return result


def _read_text_with_fallback(file_path: Path) -> str:
    """CSV 多编码兜底（utf-8 → gbk → latin-1 replace）+ BOM 剥离

    Windows 上导出的 CSV 多为 gbk，macOS 多为 utf-8-sig（带 BOM）；latin-1 永不抛
    UnicodeDecodeError（兜底），用 errors='replace' 防极端情况崩溃。

    utf-8 decode 时 BOM ('﻿') 会被当作普通字符保留，需手动剥（utf-8-sig 优先
    方案不可行：utf-8-sig 对纯 gbk 文件会 UnicodeDecodeError，顺序敏感）。
    """
    raw = file_path.read_bytes()
    for enc in ("utf-8", "gbk"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("latin-1", errors="replace")
    # 剥 BOM（utf-8 / utf-16 等都可能带）
    if text.startswith("﻿"):
        text = text[1:]
    return text


# ============ Excel 解析器 ============


class ExcelParser:
    """openpyxl 解析 .xlsx / .xls，仅读 active sheet

    max_bytes = 50MB（spec §16.1）

    用 read_only=True + data_only=True 减少内存占用（公式只取缓存值）。
    超 50MB 文件留 v1.6+ 流式优化。
    """

    mime_types = (
        "application/vnd.ms-excel",  # .xls（旧版 Office）
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    )
    max_bytes = 50 * 1024 * 1024

    async def parse(self, file_path: Path) -> FileParseResult:
        return await asyncio.to_thread(self._check_and_parse, file_path)

    def _check_and_parse(self, file_path: Path) -> FileParseResult:
        size = file_path.stat().st_size
        if size > self.max_bytes:
            raise BusinessRuleException(
                f"Excel 文件过大（{_mb(size)}MB），上限 {_mb(self.max_bytes)}MB",
                error_code="AI_FILE_TOO_LARGE",
            )
        return self._parse_sync(file_path, size)

    def _parse_sync(self, file_path: Path, size: int) -> FileParseResult:
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        try:
            sheet = wb.active
            if sheet is None:
                raise BusinessRuleException(
                    "Excel 无工作表",
                    error_code="AI_FILE_EMPTY",
                )
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return FileParseResult(
                    rows=0,
                    columns=[],
                    preview=[],
                    parser="ExcelParser",
                    file_size=size,
                )
            columns = [str(c) if c is not None else "" for c in header_row]
            preview: list[dict[str, str]] = []
            total_rows = 0
            for row in rows_iter:
                total_rows += 1
                if len(preview) < PREVIEW_ROW_LIMIT:
                    preview.append(_row_to_dict(columns, row))
            return FileParseResult(
                rows=total_rows,
                columns=columns,
                preview=preview,
                parser="ExcelParser",
                file_size=size,
            )
        finally:
            wb.close()


# ============ CSV 解析器 ============


class CsvParser:
    """csv 标准库解析（无外部依赖）

    max_bytes = 10MB（spec §16.1）
    """

    mime_types = (
        "text/csv",
        "text/plain",  # .txt / 无 MIME 时浏览器兜底，部分 .csv 落到这里
        "application/csv",  # 部分 OS 识别
    )
    max_bytes = 10 * 1024 * 1024

    async def parse(self, file_path: Path) -> FileParseResult:
        return await asyncio.to_thread(self._check_and_parse, file_path)

    def _check_and_parse(self, file_path: Path) -> FileParseResult:
        size = file_path.stat().st_size
        if size > self.max_bytes:
            raise BusinessRuleException(
                f"CSV 文件过大（{_mb(size)}MB），上限 {_mb(self.max_bytes)}MB",
                error_code="AI_FILE_TOO_LARGE",
            )
        return self._parse_sync(file_path, size)

    def _parse_sync(self, file_path: Path, size: int) -> FileParseResult:
        text = _read_text_with_fallback(file_path)
        reader = csv.reader(io.StringIO(text))
        try:
            header_row = next(reader)
        except StopIteration:
            return FileParseResult(
                rows=0,
                columns=[],
                preview=[],
                parser="CsvParser",
                file_size=size,
            )
        columns = [str(c) for c in header_row]
        preview: list[dict[str, str]] = []
        total_rows = 0
        for row in reader:
            total_rows += 1
            if len(preview) < PREVIEW_ROW_LIMIT:
                preview.append(_row_to_dict(columns, tuple(row)))
        return FileParseResult(
            rows=total_rows,
            columns=columns,
            preview=preview,
            parser="CsvParser",
            file_size=size,
        )


# ============ PARSERS 注册表 ============


def _build_parsers() -> dict[str, FileParser]:
    """构建 MIME → parser 映射，重复 MIME 启动即报错（防配置漂移）"""
    registry: dict[str, FileParser] = {}
    for parser_cls in (ExcelParser, CsvParser):
        instance = parser_cls()
        for mime in instance.mime_types:
            if mime in registry:
                existing = registry[mime].__class__.__name__
                raise RuntimeError(
                    f"MIME {mime!r} 已被 {existing} 注册，"
                    f"{parser_cls.__name__} 无法接管（请检查 mime_types 配置）"
                )
            registry[mime] = instance
    return registry


PARSERS: dict[str, FileParser] = _build_parsers()
"""MIME → parser 实例映射（启动时构建一次，运行时只读）"""

SUPPORTED_MIME_TYPES: frozenset[str] = frozenset(PARSERS.keys())
"""所有 parser 覆盖的 MIME 集合（check_ai_tools 校验 + file.parse accepts_file 用）"""


def _mb(num_bytes: int) -> str:
    """字节数 → MB 字符串（保留 1 位小数）"""
    return f"{num_bytes / 1024 / 1024:.1f}"


# ============ 入口函数 ============


async def parse_file(file_path: Path, mime_type: str) -> FileParseResult:
    """按 MIME 选 parser 解析文件

    Args:
        file_path: 文件绝对路径（由 file_tools.py 从 sys_file.file_path 拼成）
        mime_type: 文件 MIME（来自 sys_file.mime_type）

    Raises:
        BusinessRuleException: AI_FILE_TYPE_UNSUPPORTED / AI_FILE_TOO_LARGE / AI_FILE_EMPTY
    """
    parser = PARSERS.get(mime_type)
    if parser is None:
        raise BusinessRuleException(
            f"不支持的文件类型: {mime_type}",
            error_code="AI_FILE_TYPE_UNSUPPORTED",
        )
    return await parser.parse(file_path)
