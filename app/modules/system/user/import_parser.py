"""Excel/CSV 解析与字段校验。

职责：
- ``parse_import_excel(file_bytes, mime_type) -> list[UserImportRecord]``
- MIME 白名单、文件大小和行数硬上限校验
- 字段级格式校验：必填、长度、邮箱、手机号、gender、status
- 失败行一次性抛 ``ImportErrorCollection``（不一次一个）
- ``employee_no`` 空串规范化为 ``None``

**不在本模块做**：
- ``dept_input`` / ``role_input`` 存在性反查（由预览阶段处理）
- ``file_sha256`` 计算（由预览调用方处理）
- DB 查询（解析层纯函数，便于复用 + 单测）

user_service.parse_import_excel（service 层）是 thin wrapper：直接 delegate 到本模块。
"""

import csv
import io
import re
import struct
import warnings
import zipfile
from pathlib import PurePosixPath
from typing import NoReturn
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import BusinessRuleException
from app.modules.system.user.constants import USER_IMPORT_MAX_ROWS
from app.modules.system.user.schemas import FailedRow, UserImportRecord

#: 允许导入的 MIME 类型。
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_CSV = "text/csv"
ALLOWED_MIME_TYPES: frozenset[str] = frozenset({MIME_XLSX, MIME_CSV})

#: 固定 10MB 安全上限，避免运行时配置意外放宽边界。
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
XLSX_MAX_ZIP_ENTRIES = 2048
XLSX_MAX_ENTRY_UNCOMPRESSED_BYTES = 32 * 1024 * 1024
XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
XLSX_MAX_COMPRESSION_RATIO = 200.0
XLSX_MAX_DECLARED_COLUMNS = 256

_XLSX_REQUIRED_MEMBERS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
)

#: 解析支持的表头顺序；实际按表头名称匹配列索引。
EXCEL_HEADERS: tuple[str, ...] = (
    "user_name",
    "employee_no",
    "nickname",
    "user_email",
    "user_phone",
    "dept_input",
    "role_input",
    "user_gender",
    "status",
)

#: 邮箱格式（简单 RFC 5322 子集，足够滤掉常见错字）
_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
#: 中国大陆手机号格式（11 位，1 开头，第二位 3-9）
_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")

#: gender / status 合法取值（与 UserImportRecord Literal 对齐）
_GENDER_VALUES: frozenset[str] = frozenset({"0", "1", "2"})
#: status 取值与数据库和前端约定一致。
_STATUS_VALUES: frozenset[str] = frozenset({"1", "2"})

#: 中文标签反查表，使导出的中文值可再次导入。
#: 反向与 export_service._STATUS_LABELS / _GENDER_LABELS 一一对应。
#: 反查失败 fallback 到字面值继续走 _STATUS_VALUES / _GENDER_VALUES 校验。
_STATUS_LABELS_INV: dict[str, str] = {"启用": "1", "禁用": "2"}
_GENDER_LABELS_INV: dict[str, str] = {"未知": "0", "男": "1", "女": "2"}


class ImportErrorCollection(Exception):
    """一次性收集所有字段格式错误。

    一次收集所有 FailedRow（避免用户改一行重传一次）。HTTP 层 catch 后转
    400 响应，errorCode 由前端 i18n 表兜底（ ``AI_IMPORT_FIELD_ERRORS`` 或
    直接展示 ``errors[].error_code``）。

    Attributes:
        errors: 所有失败行的 FailedRow 列表（按 row_num 升序）
    """

    def __init__(self, errors: list[FailedRow]) -> None:
        self.errors = errors
        super().__init__(f"{len(errors)} field errors")


def parse_import_excel(file_bytes: bytes, mime_type: str) -> list[UserImportRecord]:
    """解析并验证 Excel/CSV，返回规范化 records。

    流程：
    1. MIME 白名单校验 → ``AI_IMPORT_INVALID_MIME``
    2. 文件大小校验 → ``AI_IMPORT_FILE_TOO_LARGE``
    3. 解析为 raw rows（xlsx 用 openpyxl / csv 用标准库）
    4. 行数硬上限校验 → ``AI_IMPORT_TOO_MANY_ROWS``
    5. 每行字段校验，累积 FailedRow
    6. 任一 FailedRow → 抛 ``ImportErrorCollection``（含全部错误）

    Args:
        file_bytes: 文件二进制内容
        mime_type: Content-Type，必须在 ALLOWED_MIME_TYPES 内

    Returns:
        list[UserImportRecord]：合法行（按 Excel 行号升序）

    Raises:
        BusinessRuleException: ``AI_IMPORT_INVALID_MIME`` / ``AI_IMPORT_FILE_TOO_LARGE``
            / ``AI_IMPORT_TOO_MANY_ROWS``
        ImportErrorCollection: 含所有字段格式 FailedRow（即使有合法行，只要有错就抛）
    """
    if mime_type not in ALLOWED_MIME_TYPES:
        raise BusinessRuleException(
            f"不支持的文件类型: {mime_type}（允许 xlsx/csv）",
            error_code="AI_IMPORT_INVALID_MIME",
        )

    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise BusinessRuleException(
            f"文件超过 {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB 限制"
            f"（实际 {len(file_bytes) / (1024 * 1024):.1f}MB）",
            error_code="AI_IMPORT_FILE_TOO_LARGE",
        )

    if mime_type == MIME_CSV:
        raw_rows = _parse_csv_rows(file_bytes)
    else:
        raw_rows = _parse_xlsx_rows(file_bytes)

    if len(raw_rows) > USER_IMPORT_MAX_ROWS:
        _raise_too_many_rows(len(raw_rows))

    records: list[UserImportRecord] = []
    errors: list[FailedRow] = []
    for idx, raw in enumerate(raw_rows, start=2):  # row 1 是表头，数据从 row 2 起
        record, row_errors = _validate_row(idx, raw)
        if record is not None:
            records.append(record)
        errors.extend(row_errors)

    if errors:
        raise ImportErrorCollection(errors)

    return records


def _parse_xlsx_rows(file_bytes: bytes) -> list[dict[str, str]]:
    """openpyxl 解析 xlsx → list[row_dict]（按 EXCEL_HEADERS 提取列）。

    - read_only + data_only 模式（不加载公式 / 样式）
    - 表头按名称匹配（大小写不敏感），允许列顺序变化
    - 完全空行跳过（防止模板尾部空白行被误判）
    - 单元格 None → ""（统一字符串处理）
    """
    _validate_xlsx_archive(file_bytes)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=("Data Validation extension is not supported and will be removed"),
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        return _load_xlsx_value_rows(file_bytes)


def _load_xlsx_value_rows(file_bytes: bytes) -> list[dict[str, str]]:
    """Read cell values while the caller owns extension-warning policy."""
    try:
        wb = load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (
        EOFError,
        IndexError,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        RuntimeError,
        SyntaxError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ):
        _raise_invalid_xlsx()
    try:
        ws = wb.active
        if ws is None:
            _raise_invalid_xlsx()
        declared_rows = ws.max_row or 0
        declared_columns = ws.max_column or 0
        if declared_rows > USER_IMPORT_MAX_ROWS + 1:
            _raise_too_many_rows(declared_rows - 1)
        if declared_columns > XLSX_MAX_DECLARED_COLUMNS:
            _raise_xlsx_expansion_too_large()
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        header_idx = _resolve_header_indices(header_row)

        result: list[dict[str, str]] = []
        for raw in rows_iter:
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            row_dict: dict[str, str] = {}
            for fname, i in header_idx.items():
                if i is None or i >= len(raw) or raw[i] is None:
                    row_dict[fname] = ""
                else:
                    row_dict[fname] = str(raw[i]).strip()
            result.append(row_dict)
            if len(result) > USER_IMPORT_MAX_ROWS:
                _raise_too_many_rows(len(result))
        return result
    except (
        EOFError,
        IndexError,
        InvalidFileException,
        KeyError,
        OSError,
        ParseError,
        RuntimeError,
        SyntaxError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
    ):
        _raise_invalid_xlsx()
    finally:
        wb.close()


def _parse_csv_rows(file_bytes: bytes) -> list[dict[str, str]]:
    """csv 标准库解析 → list[row_dict]。

    - utf-8-sig 解码（兼容 Excel 导出的 BOM）
    - 表头大小写不敏感匹配
    - 完全空行跳过
    """
    text = file_bytes.decode("utf-8-sig", errors="strict")
    line_separators = text.count("\n") + text.count("\r") - text.count("\r\n")
    physical_lines = line_separators + bool(text and not text.endswith(("\n", "\r")))
    if physical_lines > USER_IMPORT_MAX_ROWS + 1:
        _raise_too_many_rows(physical_lines - 1)
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    header_idx = _resolve_header_indices(fieldnames)

    result: list[dict[str, str]] = []
    for raw in reader:
        if all((v or "").strip() == "" for v in raw.values()):
            continue
        row_dict: dict[str, str] = {}
        for fname, _ in header_idx.items():
            # DictReader 用原始 fieldnames 索引，需要回查
            csv_key = _find_csv_key_for_field(fieldnames, fname)
            cell = raw.get(csv_key) if csv_key else None
            row_dict[fname] = "" if cell is None else str(cell).strip()
        result.append(row_dict)
        if len(result) > USER_IMPORT_MAX_ROWS:
            _raise_too_many_rows(len(result))
    return result


def _raise_too_many_rows(actual_rows: int) -> NoReturn:
    raise BusinessRuleException(
        f"行数超 {USER_IMPORT_MAX_ROWS} 上限（实际至少 {actual_rows} 行），"
        "请拆分为多个文件后分批导入",
        error_code="AI_IMPORT_TOO_MANY_ROWS",
    )


def _raise_invalid_xlsx() -> NoReturn:
    raise BusinessRuleException(
        "文件不是有效的 XLSX 工作簿",
        error_code="AI_IMPORT_INVALID_MIME",
    )


def _raise_xlsx_expansion_too_large() -> NoReturn:
    raise BusinessRuleException(
        "XLSX 解压结构超过安全预算",
        error_code="AI_IMPORT_FILE_TOO_LARGE",
    )


def _safe_zip_member_name(name: str) -> bool:
    if not name or "\x00" in name or "\\" in name:
        return False
    path = PurePosixPath(name)
    return (
        not path.is_absolute()
        and ".." not in path.parts
        and not (path.parts and ":" in path.parts[0])
    )


def _declared_zip_entry_count(file_bytes: bytes) -> int:
    """Read EOCD before ZipFile allocates one ZipInfo per declared member."""
    signature = b"PK\x05\x06"
    search_start = max(0, len(file_bytes) - (65_535 + 22))
    search_end = len(file_bytes)
    while True:
        offset = file_bytes.rfind(signature, search_start, search_end)
        if offset < 0:
            _raise_invalid_xlsx()
        if offset + 22 <= len(file_bytes):
            (
                disk_number,
                central_directory_disk,
                entries_on_disk,
                total_entries,
                central_directory_size,
                central_directory_offset,
                comment_length,
            ) = struct.unpack_from("<4H2LH", file_bytes, offset + 4)
            if offset + 22 + comment_length == len(file_bytes):
                if (
                    disk_number != 0
                    or central_directory_disk != 0
                    or entries_on_disk != total_entries
                ):
                    _raise_invalid_xlsx()
                if (
                    total_entries == 0xFFFF
                    or central_directory_size == 0xFFFFFFFF
                    or central_directory_offset == 0xFFFFFFFF
                ):
                    _raise_xlsx_expansion_too_large()
                if total_entries > XLSX_MAX_ZIP_ENTRIES:
                    _raise_xlsx_expansion_too_large()
                if central_directory_offset + central_directory_size != offset:
                    _raise_invalid_xlsx()
                return total_entries
        search_end = offset


def _validate_xlsx_archive(file_bytes: bytes) -> None:
    """Bound ZIP expansion before openpyxl sees attacker-controlled OOXML."""
    if not file_bytes.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        _raise_invalid_xlsx()
    declared_entries = _declared_zip_entry_count(file_bytes)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            infos = archive.infolist()
            if len(infos) != declared_entries:
                _raise_invalid_xlsx()
            if len(infos) > XLSX_MAX_ZIP_ENTRIES:
                _raise_xlsx_expansion_too_large()

            names: set[str] = set()
            total_uncompressed = 0
            total_compressed = 0
            for info in infos:
                is_symlink = (info.external_attr >> 16) & 0o170000 == 0o120000
                if (
                    not _safe_zip_member_name(info.filename)
                    or info.filename in names
                    or info.flag_bits & 0x1
                    or is_symlink
                ):
                    _raise_invalid_xlsx()
                names.add(info.filename)

                if info.file_size > XLSX_MAX_ENTRY_UNCOMPRESSED_BYTES:
                    _raise_xlsx_expansion_too_large()
                if (
                    info.file_size > 0
                    and info.file_size / max(1, info.compress_size)
                    > XLSX_MAX_COMPRESSION_RATIO
                ):
                    _raise_xlsx_expansion_too_large()
                total_uncompressed += info.file_size
                total_compressed += info.compress_size

            if total_uncompressed > XLSX_MAX_TOTAL_UNCOMPRESSED_BYTES or (
                total_uncompressed > 0
                and total_uncompressed / max(1, total_compressed)
                > XLSX_MAX_COMPRESSION_RATIO
            ):
                _raise_xlsx_expansion_too_large()

            if not _XLSX_REQUIRED_MEMBERS.issubset(names) or not any(
                name.startswith("xl/worksheets/") and name.endswith(".xml")
                for name in names
            ):
                _raise_invalid_xlsx()
            if archive.testzip() is not None:
                _raise_invalid_xlsx()
    except BusinessRuleException:
        raise
    except (
        EOFError,
        KeyError,
        NotImplementedError,
        OSError,
        RuntimeError,
        ValueError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
    ):
        _raise_invalid_xlsx()


def _resolve_header_indices(header_row) -> dict[str, int | None]:
    """表头行 → {field_name: col_index | None}。

    表头大小写不敏感匹配；缺失字段的 col_index = None，_validate_row 视为空值。
    """
    lowered: dict[str, int] = {}
    for i, c in enumerate(header_row):
        if c is None:
            continue
        lowered.setdefault(str(c).strip().lower(), i)
    return {fname: lowered.get(fname.lower()) for fname in EXCEL_HEADERS}


def _find_csv_key_for_field(fieldnames: list[str], fname: str) -> str | None:
    """DictReader 用原始 key 取值，需要大小写不敏感回查。"""
    for fn in fieldnames:
        if fn and fn.strip().lower() == fname.lower():
            return fn
    return None


def _validate_row(
    row_num: int, raw: dict[str, str]
) -> tuple[UserImportRecord | None, list[FailedRow]]:
    """单行 → (record | None, [failed_rows])。

    校验顺序：必填 → 长度 → 格式 → 取值。一行可产生多个 FailedRow。
    任一错误 → record 为 None（保证返回的 records 都是合法的）。
    """
    errors: list[FailedRow] = []

    def _get(name: str) -> str:
        v = raw.get(name)
        return "" if v is None else str(v).strip()

    # user_name 必填 + 长度 2-16（统一用 AI_IMPORT_USERNAME_INVALID）
    user_name = _get("user_name")
    if not user_name:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="user_name",
                value=user_name,
                reason="user_name 必填",
                error_code="AI_IMPORT_USERNAME_INVALID",
            )
        )
    elif len(user_name) < 2 or len(user_name) > 16:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="user_name",
                value=user_name,
                reason="user_name 长度需 2-16 字符",
                error_code="AI_IMPORT_USERNAME_INVALID",
            )
        )

    # employee_no 可选，最长 64；空串规范化为 None。
    employee_no_raw = _get("employee_no")
    employee_no: str | None = employee_no_raw or None
    if employee_no and len(employee_no) > 64:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="employee_no",
                value=employee_no,
                reason="employee_no 长度超 64",
                error_code="AI_IMPORT_EMPLOYEE_NO_TOO_LONG",
            )
        )

    # nickname 可选，max 16
    nickname_raw = _get("nickname")
    nickname: str | None = nickname_raw or None
    if nickname and len(nickname) > 16:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="nickname",
                value=nickname,
                reason="nickname 长度超 16",
                error_code="AI_IMPORT_NICKNAME_TOO_LONG",
            )
        )

    # user_email 可选，max 128，邮箱格式
    email_raw = _get("user_email")
    email: str | None = email_raw or None
    if email:
        if len(email) > 128:
            errors.append(
                FailedRow(
                    row_num=row_num,
                    field="user_email",
                    value=email,
                    reason="user_email 长度超 128",
                    error_code="AI_IMPORT_EMAIL_INVALID",
                )
            )
        elif not _EMAIL_RE.match(email):
            errors.append(
                FailedRow(
                    row_num=row_num,
                    field="user_email",
                    value=email,
                    reason="邮箱格式错",
                    error_code="AI_IMPORT_EMAIL_INVALID",
                )
            )

    # user_phone 可选，max 32，中国大陆手机号格式
    phone_raw = _get("user_phone")
    phone: str | None = phone_raw or None
    if phone:
        if len(phone) > 32:
            errors.append(
                FailedRow(
                    row_num=row_num,
                    field="user_phone",
                    value=phone,
                    reason="user_phone 长度超 32",
                    error_code="AI_IMPORT_PHONE_INVALID",
                )
            )
        elif not _PHONE_RE.match(phone):
            errors.append(
                FailedRow(
                    row_num=row_num,
                    field="user_phone",
                    value=phone,
                    reason="手机号格式错（11 位数字，1 开头）",
                    error_code="AI_IMPORT_PHONE_INVALID",
                )
            )

    # dept_input 必填
    dept_input = _get("dept_input")
    if not dept_input:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="dept_input",
                value=dept_input,
                reason="dept_input 必填（部门名或完整路径）",
                error_code="AI_IMPORT_DEPT_INPUT_REQUIRED",
            )
        )

    # role_input 可选
    role_input_raw = _get("role_input")
    role_input: str | None = role_input_raw or None

    # user_gender Literal["0","1","2"]，默认 "0"
    # 先反查中文标签，未命中时按原始枚举值处理。
    gender_input = _get("user_gender") or "0"
    gender_raw = _GENDER_LABELS_INV.get(gender_input, gender_input)
    if gender_raw not in _GENDER_VALUES:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="user_gender",
                value=gender_input,
                reason="user_gender 取值需 0/1/2 或 未知/男/女",
                error_code="AI_IMPORT_GENDER_INVALID",
            )
        )
        gender = "0"
    else:
        gender = gender_raw

    # status 取值为 "1"/"2"，默认 "1"；中文标签可反向解析。
    status_input = _get("status") or "1"
    status_raw = _STATUS_LABELS_INV.get(status_input, status_input)
    if status_raw not in _STATUS_VALUES:
        errors.append(
            FailedRow(
                row_num=row_num,
                field="status",
                value=status_input,
                reason="status 取值需 1/2 或 启用/禁用",
                error_code="AI_IMPORT_STATUS_INVALID",
            )
        )
        status = "1"
    else:
        status = status_raw

    if errors:
        return None, errors

    record = UserImportRecord(
        row_num=row_num,
        user_name=user_name,
        employee_no=employee_no,
        nickname=nickname,
        user_email=email,
        user_phone=phone,
        dept_input=dept_input,
        role_input=role_input,
        user_gender=gender,
        status=status,
    )
    return record, []


__all__ = [
    "ALLOWED_MIME_TYPES",
    "EXCEL_HEADERS",
    "MAX_FILE_SIZE_BYTES",
    "MIME_CSV",
    "MIME_XLSX",
    "ImportErrorCollection",
    "parse_import_excel",
]
