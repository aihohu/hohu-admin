from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import cacheable
from app.core.exceptions import DuplicateException, NotFoundException
from app.modules.system.models.config import Config
from app.modules.system.schemas.config import (
    ConfigCreate,
    ConfigQuery,
    ConfigUpdate,
)
from app.utils.pagination import build_filters, paginate

EXCEL_HEADERS = [
    "config_name",
    "config_key",
    "config_value",
    "config_type",
    "config_group",
    "status",
    "is_public",
    "remark",
]


class ConfigService:
    """系统配置业务逻辑服务"""

    async def get_list(self, db: AsyncSession, query: ConfigQuery):
        """获取系统配置分页列表"""
        field_mapping = {
            "config_name": ("config_name", "contains"),
            "config_key": ("config_key", "contains"),
            "config_group": ("config_group", "contains"),
            "status": ("status", "=="),
        }
        filters = build_filters(Config, field_mapping, **query.model_dump())

        page_data = await paginate(
            db=db,
            model=Config,
            query_params=query,
            filters=filters,
            order_by=Config.config_group.asc().nulls_last(),
        )

        return page_data

    @cacheable(key="config:public", ttl=300)
    async def get_public_configs(self, db: AsyncSession) -> dict[str, str]:
        """获取公开配置（无需鉴权），返回 {key: value} 字典"""
        result = await db.execute(
            select(Config)
            .where(Config.is_public == True, Config.status == "1")  # noqa: E712
            .order_by(Config.config_group.asc(), Config.config_key.asc())
        )
        return {c.config_key: c.config_value for c in result.scalars().all()}

    @cacheable(key="config:key:{key}", ttl=300)
    async def get_value(
        self, db: AsyncSession, key: str, default: str | None = None
    ) -> str | None:
        """根据 key 获取配置值，支持默认值"""
        result = await db.execute(
            select(Config.config_value).where(
                Config.config_key == key, Config.status == "1"
            )
        )
        value = result.scalar_one_or_none()
        return value if value is not None else default

    @cacheable(key="config:group:{group}", ttl=300)
    async def get_values_by_group(self, db: AsyncSession, group: str) -> dict[str, str]:
        """根据分组获取配置，返回 {key: value} 字典"""
        result = await db.execute(
            select(Config)
            .where(Config.config_group == group, Config.status == "1")
            .order_by(Config.config_key.asc())
        )
        return {c.config_key: c.config_value for c in result.scalars().all()}

    async def create(self, db: AsyncSession, config_in: ConfigCreate) -> Config:
        """创建系统配置"""
        # 检查键唯一性
        check = await db.execute(
            select(Config).where(Config.config_key == config_in.config_key)
        )
        if check.scalars().first():
            raise DuplicateException("配置键", config_in.config_key)

        new_config = Config(**config_in.model_dump())
        db.add(new_config)
        return new_config

    async def update(
        self, db: AsyncSession, config_id: int, config_in: ConfigUpdate
    ) -> Config:
        """更新系统配置"""
        config = await db.get(Config, config_id)
        if not config:
            raise NotFoundException("系统配置")

        # 如果修改了 config_key，检查唯一性
        update_data = config_in.model_dump(exclude_unset=True)
        if (
            "config_key" in update_data
            and update_data["config_key"] != config.config_key
        ):
            check = await db.execute(
                select(Config).where(Config.config_key == update_data["config_key"])
            )
            if check.scalars().first():
                raise DuplicateException("配置键", update_data["config_key"])

        for field, value in update_data.items():
            setattr(config, field, value)

        return config

    async def delete(self, db: AsyncSession, config_id: int) -> None:
        """删除系统配置"""
        config = await db.get(Config, config_id)
        if not config:
            raise NotFoundException("系统配置")

        await db.delete(config)

    async def batch_delete(self, db: AsyncSession, ids: list[int]) -> int:
        """批量删除系统配置"""
        result = await db.execute(select(Config).where(Config.config_id.in_(ids)))
        config_list = result.scalars().all()

        for config in config_list:
            await db.delete(config)

        return len(config_list)

    async def export_configs(self, db: AsyncSession, query: ConfigQuery) -> BytesIO:
        """导出系统配置为 Excel"""
        field_mapping = {
            "config_name": ("config_name", "contains"),
            "config_key": ("config_key", "contains"),
            "config_group": ("config_group", "contains"),
            "status": ("status", "=="),
        }
        filters = build_filters(Config, field_mapping, **query.model_dump())
        stmt = select(Config).where(*filters) if filters else select(Config)
        result = await db.execute(
            stmt.order_by(Config.config_group.asc(), Config.config_key.asc())
        )
        configs = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "系统配置"
        ws.append(EXCEL_HEADERS)
        for c in configs:
            ws.append(
                [
                    c.config_name,
                    c.config_key,
                    c.config_value,
                    c.config_type,
                    c.config_group,
                    c.status,
                    c.is_public,
                    c.remark,
                ]
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def import_configs(
        self, db: AsyncSession, file_bytes: bytes
    ) -> dict[str, int]:
        """从 Excel 导入系统配置，返回 {success: n, skipped: n}"""
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        # 查询已存在的 key
        existing = await db.execute(select(Config.config_key))
        existing_keys = set(existing.scalars().all())

        success = 0
        skipped = 0
        for row in rows:
            if not row or len(row) < 2 or not row[1]:
                continue
            config_key = str(row[1]).strip()
            if config_key in existing_keys:
                skipped += 1
                continue

            config = Config(
                config_name=str(row[0] or ""),
                config_key=config_key,
                config_value=str(row[2] or ""),
                config_type=str(row[3] or "text"),
                config_group=str(row[4] or ""),
                status=str(row[5] or "1"),
                is_public=bool(row[6]) if row[6] is not None else False,
                remark=str(row[7] or "") if len(row) > 7 and row[7] else None,
            )
            db.add(config)
            existing_keys.add(config_key)
            success += 1

        return {"success": success, "skipped": skipped}


# 创建单例
config_service = ConfigService()
