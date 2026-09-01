"""用户导入模板生成服务。

职责：
- 生成 4 sheet xlsx：「数据」/「说明」/「部门字典」/「角色字典」
- 「数据」sheet 列顺序固定并提供两行示例
- 「数据」sheet E 列（dept_input）+ F 列（role_input）加 DataValidation
  下拉，引用字典 sheet
- 「部门字典」/「角色字典」sheet 实时查询并标注生成时间

不在本模块做：
- HTTP 响应封装（API 层用 Response(content=bytes) 包装）
- 权限校验（API 层 require_permissions）
- 缓存（模板每次下载都重新生成，字典 sheet 数据实时）
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.tenant import TenantContext
from app.modules.system.models.dept import Dept
from app.modules.system.models.role import Role

#: 「数据」sheet 列顺序。
#: 与 ``import_parser.EXCEL_HEADERS`` 不完全一致：模板展示 8 列，
#: 8 列展示（不带 employee_no），但 parser 表头大小写不敏感匹配可容忍缺失列。
_DATA_COLUMNS: tuple[str, ...] = (
    "user_name",
    "nickname",
    "user_email",
    "user_phone",
    "dept_input",
    "role_input",
    "user_gender",
    "status",
)

#: 两行示例帮助用户理解必填项和格式。
#: 示例值用明显的「示例」语义，避免用户直接落库（dry_run 会拦下重名）。
_DATA_EXAMPLE_ROWS: tuple[tuple[str, ...], ...] = (
    (
        "zhangsan",
        "张三",
        "zhangsan@example.com",
        "13800138000",
        "总公司/研发中心/前端部",
        "R_DEV",
        "1",
        "1",
    ),
    (
        "lisi",
        "李四",
        "lisi@example.com",
        "13900139000",
        "总公司/市场部",
        "R_OPS,内容编辑",
        "0",
        "1",
    ),
)

#: 「说明」sheet 字段说明。
#: 每行 4 列：字段名 / 必填 / 取值范围 / 冲突处理。
_INSTRUCTION_ROWS: tuple[tuple[str, str, str, str], ...] = (
    ("字段名", "必填", "取值范围", "冲突处理策略"),
    (
        "user_name",
        "是",
        "2-16 字符，字母数字下划线",
        "重名按 on_conflict 策略（skip / overwrite / fail_fast）",
    ),
    (
        "nickname",
        "否",
        "≤ 16 字符",
        "—",
    ),
    (
        "user_email",
        "否",
        "RFC 邮箱格式，≤ 128 字符",
        "重名按 on_conflict 策略",
    ),
    (
        "user_phone",
        "否",
        "11 位中国手机号，1 开头",
        "重名按 on_conflict 策略",
    ),
    (
        "dept_input",
        "是",
        "部门名（如「前端部」）或完整路径（如「总公司/研发中心/前端部」）",
        "查不到 → AI_IMPORT_DEPT_NOT_FOUND；重名 → AI_IMPORT_DEPT_DUPLICATE",
    ),
    (
        "role_input",
        "否",
        "逗号分隔，role_code 或 role_name 混合（如「R_DEV,内容编辑」）",
        "查不到 → AI_IMPORT_ROLE_NOT_FOUND；越权 → AI_IMPORT_ROLE_OUT_OF_SCOPE",
    ),
    (
        "user_gender",
        "否（默认 0）",
        "0 未知 / 1 男 / 2 女",
        "—",
    ),
    (
        "status",
        "否（默认 1）",
        "1 启用 / 2 禁用",
        "—",
    ),
)

#: 字典 sheet 表头。
_DEPT_DICT_COLUMNS: tuple[str, ...] = (
    "dept_name",
    "full_path",
    "dept_id",
    "status",
)
_ROLE_DICT_COLUMNS: tuple[str, ...] = (
    "role_code",
    "role_name",
    "status",
)

#: 顶部生成时间标注。
_TIMESTAMP_LABEL = "⏰ 生成时间"
_TIMESTAMP_HINT = "（数据可能已变化，请重新下载模板获取最新）"

#: 数据验证应用范围。
_DEPT_DV_RANGE = "E2:E10000"
_ROLE_DV_RANGE = "F2:F10000"

#: 字典 sheet 顶部 row 1 是生成时间标注，row 2 是表头，
#: row 3+ 才是数据行 —— DataValidation formula 起点是 ``$X$3`` 而非 ``$X$2``。
_DEPT_DV_FORMULA = "部门字典!$B$3:$B$1000"
_ROLE_DV_FORMULA = "角色字典!$A$3:$A$50"


async def _fetch_depts(db: AsyncSession, *, tenant: TenantContext) -> list[Dept]:
    """实时查询启用部门。

    - 仅 status='1' 启用部门（决策：禁用部门不展示给用户选）
    - 按 dept_id 升序（稳定输出，便于测试 + 用户对比）
    """
    stmt = (
        select(Dept)
        .where(
            Dept.tenant_id == tenant.tenant_id,
            Dept.status == "1",
        )
        .order_by(Dept.dept_id.asc())
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _fetch_roles(db: AsyncSession, *, tenant: TenantContext) -> list[Role]:
    """实时查询启用角色。

    - 仅 status='1' 启用角色
    - 按 role_id 升序
    """
    stmt = (
        select(Role)
        .where(
            Role.tenant_id == tenant.tenant_id,
            Role.status == "1",
        )
        .order_by(Role.role_id.asc())
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


def _build_dept_full_path(dept: Dept, dept_lookup: dict[int, Dept]) -> str:
    """根据 ancestors 链构建 full_path（如「总公司/研发中心/前端部」）。

    ancestors 字段格式：``"0,1234,5678"``（顶级部门 ancestors='0'）。
    - 解析 ancestors → 拿到祖先 dept_id 列表
    - 按 ancestors 顺序拼接 dept_name + 当前 dept_name

    Args:
        dept: 当前部门
        dept_lookup: dept_id → Dept 映射（含祖先）

    Returns:
        ``"祖先1/祖先2/.../dept_name"``；祖先查不到时回退到 dept_name 本身。
    """
    if not dept.ancestors:
        return dept.dept_name

    parts: list[str] = []
    for raw in dept.ancestors.split(","):
        raw = raw.strip()
        if not raw or raw == "0":
            continue
        try:
            ancestor_id = int(raw)
        except ValueError:
            continue
        ancestor = dept_lookup.get(ancestor_id)
        if ancestor is not None:
            parts.append(ancestor.dept_name)

    parts.append(dept.dept_name)
    return "/".join(parts)


def _build_data_sheet(ws) -> None:
    """「数据」sheet：表头 + 2 行示例 + DataValidation。

    包含固定表头、示例行和字典数据验证。
    """
    # 表头
    ws.append(list(_DATA_COLUMNS))

    # 两行示例。
    for row in _DATA_EXAMPLE_ROWS:
        ws.append(list(row))

    # 部门与角色数据验证。
    # 部门列（E）：引用「部门字典」sheet 的 full_path 列
    dept_dv = DataValidation(
        type="list",
        formula1=f"={_DEPT_DV_FORMULA}",
        allow_blank=False,
        showErrorMessage=True,
    )
    dept_dv.error = "请从下拉选择部门（或到「部门字典」sheet 复制 full_path）"
    dept_dv.errorTitle = "部门无效"
    ws.add_data_validation(dept_dv)
    dept_dv.add(_DEPT_DV_RANGE)

    # 角色列（F）：引用「角色字典」sheet 的 role_code 列（allow_blank=True）
    role_dv = DataValidation(
        type="list",
        formula1=f"={_ROLE_DV_FORMULA}",
        allow_blank=True,
        showErrorMessage=True,
    )
    role_dv.error = "请从下拉选择角色（或到「角色字典」sheet 复制 role_code）"
    role_dv.errorTitle = "角色无效"
    ws.add_data_validation(role_dv)
    role_dv.add(_ROLE_DV_RANGE)


def _build_instruction_sheet(ws) -> None:
    """构建字段说明、必填项、取值范围和冲突处理说明。"""
    for row in _INSTRUCTION_ROWS:
        ws.append(list(row))


def _build_dept_dict_sheet(ws, depts: list[Dept], dept_lookup: dict[int, Dept]) -> None:
    """「部门字典」sheet：生成时间行 + 表头 + 数据行。

    数据从当前启用部门实时生成。
    """
    # 顶部生成时间标注。
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"{_TIMESTAMP_LABEL}：{now_str}{_TIMESTAMP_HINT}"])

    # 表头
    ws.append(list(_DEPT_DICT_COLUMNS))

    # 数据行
    for dept in depts:
        full_path = _build_dept_full_path(dept, dept_lookup)
        ws.append(
            [
                dept.dept_name,
                full_path,
                str(dept.dept_id),
                dept.status,
            ]
        )


def _build_role_dict_sheet(ws, roles: list[Role]) -> None:
    """「角色字典」sheet：生成时间行 + 表头 + 数据行。

    数据从当前启用角色实时生成。
    """
    # 顶部生成时间标注
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append([f"{_TIMESTAMP_LABEL}：{now_str}{_TIMESTAMP_HINT}"])

    # 表头
    ws.append(list(_ROLE_DICT_COLUMNS))

    # 数据行
    for role in roles:
        ws.append(
            [
                role.role_code,
                role.role_name,
                role.status,
            ]
        )


async def generate_import_template(db: AsyncSession, *, tenant: TenantContext) -> bytes:
    """生成用户导入模板 xlsx。

    流程：
    1. 实时查询启用部门和角色
    2. 构建 dept_lookup（ancestors 反查 full_path 用）
    3. 新建 Workbook + 4 sheet（顺序：数据 / 说明 / 部门字典 / 角色字典）
    4. 各 sheet 填内容
    5. 保存为 bytes 返回

    Args:
        db: 异步数据库会话（查 sys_dept / sys_role）

    Returns:
        xlsx bytes（API 层用 ``Response(content=bytes)`` 包装）
    """
    depts = await _fetch_depts(db, tenant=tenant)
    roles = await _fetch_roles(db, tenant=tenant)
    dept_lookup: dict[int, Dept] = {d.dept_id: d for d in depts}

    wb = Workbook()
    # wb.active 默认创建第 1 个 sheet，重命名为「数据」
    ws_data = wb.active
    ws_data.title = "数据"
    _build_data_sheet(ws_data)

    # 其余 sheet 用 create_sheet（按顺序追加到末尾）
    ws_instruction = wb.create_sheet("说明")
    _build_instruction_sheet(ws_instruction)

    ws_dept = wb.create_sheet("部门字典")
    _build_dept_dict_sheet(ws_dept, depts, dept_lookup)

    ws_role = wb.create_sheet("角色字典")
    _build_role_dict_sheet(ws_role, roles)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["generate_import_template"]
