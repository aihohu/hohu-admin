"""``parse_import_excel`` 行为测试。

覆盖：
- MIME 白名单（xlsx/csv；legacy xls fail-closed）→ AI_IMPORT_INVALID_MIME
- 文件大小 ≤ 10MB → AI_IMPORT_FILE_TOO_LARGE
- 行数 ≤ 2000 → AI_IMPORT_TOO_MANY_ROWS
- 字段校验：user_name 必填 / 邮箱 / 手机 / gender / status / 长度
- 使用 ImportErrorCollection 一次性收集失败行
- 空字符串 employee_no 规范化为 None
- 可选字段默认值（gender="0" / status="1"）

不依赖 DB（解析层零 DB 查询，dept_input / role_input 存在性留给 dry_run）。
"""

import io
import warnings
import zipfile

import pytest
from openpyxl import Workbook

from app.core.exceptions import BusinessRuleException
from app.modules.system.user.constants import USER_IMPORT_MAX_ROWS
from app.modules.system.user.import_parser import (
    ALLOWED_MIME_TYPES,
    EXCEL_HEADERS,
    MAX_FILE_SIZE_BYTES,
    ImportErrorCollection,
    _parse_csv_rows,
    _parse_xlsx_rows,
    parse_import_excel,
)
from app.modules.system.user.schemas import FailedRow, UserImportRecord


def _xlsx_bytes(rows: list[list[str]], headers: list[str] | None = None) -> bytes:
    """构造 xlsx 文件 bytes（含表头 + 数据行）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    if headers is None:
        headers = [
            "user_name",
            "employee_no",
            "nickname",
            "user_email",
            "user_phone",
            "dept_input",
            "role_input",
            "user_gender",
            "status",
        ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(rows: list[list[str]], headers: list[str] | None = None) -> bytes:
    """构造 CSV 文件 bytes。"""
    if headers is None:
        headers = [
            "user_name",
            "employee_no",
            "nickname",
            "user_email",
            "user_phone",
            "dept_input",
            "role_input",
            "user_gender",
            "status",
        ]
    lines = [",".join(headers)]
    for row in rows:
        lines.append(",".join(row))
    return "\n".join(lines).encode("utf-8")


def _replace_zip_member(data: bytes, name: str, replacement: bytes) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(data))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            target.writestr(
                info,
                replacement if info.filename == name else source.read(info.filename),
            )
    return output.getvalue()


def _xlsx_with_data_validation_extension(rows: list[list[str]]) -> bytes:
    data = _xlsx_bytes(rows)
    with zipfile.ZipFile(io.BytesIO(data)) as source:
        worksheet = source.read("xl/worksheets/sheet1.xml")
    extension = (
        b'<extLst><ext uri="{CCE6A557-97BC-4B89-ADB6-D9C93CAAB3DF}">'
        b'<x14:dataValidations xmlns:x14="http://schemas.microsoft.com/'
        b'office/spreadsheetml/2009/9/main" count="0"/>'
        b"</ext></extLst>"
    )
    worksheet = worksheet.replace(b"</worksheet>", extension + b"</worksheet>")
    return _replace_zip_member(data, "xl/worksheets/sheet1.xml", worksheet)


MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_CSV = "text/csv"
MIME_INVALID = "application/pdf"


class TestMimeTypeWhitelist:
    """安全边界仅声明当前解析器真正支持的 {xlsx, csv}。"""

    def test_allowed_mime_types_include_xlsx_csv_but_not_legacy_xls(self):
        assert MIME_XLSX in ALLOWED_MIME_TYPES
        assert "application/vnd.ms-excel" not in ALLOWED_MIME_TYPES
        assert MIME_CSV in ALLOWED_MIME_TYPES

    def test_legacy_xls_is_fail_closed(self):
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(
                b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1",
                "application/vnd.ms-excel",
            )
        assert exc.value.error_code == "AI_IMPORT_INVALID_MIME"

    def test_invalid_mime_raises(self):
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(b"whatever", MIME_INVALID)
        assert exc.value.error_code == "AI_IMPORT_INVALID_MIME"


class TestFileSizeLimit:
    """文件大小上限为 10MB。"""

    def test_max_file_size_is_10mb(self):
        assert MAX_FILE_SIZE_BYTES == 10 * 1024 * 1024

    def test_too_large_raises(self):
        # 攒 10MB + 1 byte
        bogus = b"x" * (MAX_FILE_SIZE_BYTES + 1)
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(bogus, MIME_XLSX)
        assert exc.value.error_code == "AI_IMPORT_FILE_TOO_LARGE"

    def test_damaged_xlsx_has_stable_type_error(self):
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(b"PK\x03\x04damaged", MIME_XLSX)
        assert exc.value.error_code == "AI_IMPORT_INVALID_MIME"

    def test_malformed_worksheet_xml_has_stable_type_error(self):
        malformed = _replace_zip_member(
            _xlsx_bytes([["alice", "", "", "", "", "QA", "", "0", "1"]]),
            "xl/worksheets/sheet1.xml",
            b"<broken",
        )
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(malformed, MIME_XLSX)
        assert exc.value.error_code == "AI_IMPORT_INVALID_MIME"

    def test_xlsx_expansion_budget_is_enforced(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        monkeypatch.setattr(
            "app.modules.system.user.import_parser.XLSX_MAX_COMPRESSION_RATIO", 1.0
        )
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(_xlsx_bytes([]), MIME_XLSX)
        assert exc.value.error_code == "AI_IMPORT_FILE_TOO_LARGE"


class TestRowCountLimit:
    """导入行数上限为 2000。"""

    def test_user_import_max_rows_is_2000(self):
        assert USER_IMPORT_MAX_ROWS == 2000

    def test_too_many_rows_raises(self):
        rows = [
            [f"user{i}", "", f"nick{i}", "", "", "QA-Dept", "", "0", "1"]
            for i in range(USER_IMPORT_MAX_ROWS + 1)
        ]
        with pytest.raises(BusinessRuleException) as exc:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert exc.value.error_code == "AI_IMPORT_TOO_MANY_ROWS"
        assert "分批" in exc.value.message
        assert "异步" not in exc.value.message

    def test_xlsx_parser_stops_on_row_2001(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        class WorkbookStub:
            closed = False
            max_row = 1
            max_column = len(EXCEL_HEADERS)

            @property
            def active(self):
                return self

            def iter_rows(self, *, values_only: bool):
                assert values_only is True
                yield EXCEL_HEADERS
                valid = ("alice", "", "", "", "", "QA", "", "0", "1")
                for _ in range(USER_IMPORT_MAX_ROWS + 1):
                    yield valid
                raise AssertionError("parser consumed beyond row 2001")

            def close(self):
                self.closed = True

        workbook = WorkbookStub()
        monkeypatch.setattr(
            "app.modules.system.user.import_parser.load_workbook",
            lambda *_args, **_kwargs: workbook,
        )
        monkeypatch.setattr(
            "app.modules.system.user.import_parser._validate_xlsx_archive",
            lambda *_args, **_kwargs: None,
        )

        with pytest.raises(BusinessRuleException) as exc:
            _parse_xlsx_rows(b"test")

        assert exc.value.error_code == "AI_IMPORT_TOO_MANY_ROWS"
        assert workbook.closed is True

    def test_xlsx_declared_dimension_is_rejected_before_iteration(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        class WorkbookStub:
            closed = False
            max_row = 1_048_576
            max_column = 16_384

            @property
            def active(self):
                return self

            def iter_rows(self, *, values_only: bool):
                assert values_only is True
                raise AssertionError("oversized worksheet must not be iterated")

            def close(self):
                self.closed = True

        workbook = WorkbookStub()
        monkeypatch.setattr(
            "app.modules.system.user.import_parser.load_workbook",
            lambda *_args, **_kwargs: workbook,
        )
        monkeypatch.setattr(
            "app.modules.system.user.import_parser._validate_xlsx_archive",
            lambda *_args, **_kwargs: None,
        )

        with pytest.raises(BusinessRuleException) as exc:
            _parse_xlsx_rows(b"test")

        assert exc.value.error_code == "AI_IMPORT_TOO_MANY_ROWS"
        assert workbook.closed is True

    def test_csv_parser_stops_on_row_2001(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        class ReaderStub:
            fieldnames = list(EXCEL_HEADERS)

            def __iter__(self):
                valid = {
                    "user_name": "alice",
                    "dept_input": "QA",
                    "user_gender": "0",
                    "status": "1",
                }
                for _ in range(USER_IMPORT_MAX_ROWS + 1):
                    yield valid
                raise AssertionError("parser consumed beyond row 2001")

        monkeypatch.setattr(
            "app.modules.system.user.import_parser.csv.DictReader",
            lambda *_args, **_kwargs: ReaderStub(),
        )

        with pytest.raises(BusinessRuleException) as exc:
            _parse_csv_rows(b"test")

        assert exc.value.error_code == "AI_IMPORT_TOO_MANY_ROWS"


class TestParseXlsxBasic:
    """spec line 2624：标准 xlsx → records 数正确。"""

    def test_parse_xlsx_basic(self):
        rows = [
            [
                "alice",
                "E001",
                "Alice",
                "alice@example.com",
                "13800138000",
                "QA-Dept",
                "R_DEV",
                "1",
                "1",
            ],
            [
                "bob",
                "E002",
                "Bob",
                "bob@example.com",
                "13900139000",
                "QA-Dept",
                "",
                "0",
                "1",
            ],
        ]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert len(records) == 2
        assert isinstance(records[0], UserImportRecord)
        assert records[0].row_num == 2  # 表头是 row 1
        assert records[0].user_name == "alice"
        assert records[0].employee_no == "E001"
        assert records[0].nickname == "Alice"
        assert records[0].user_email == "alice@example.com"
        assert records[0].user_phone == "13800138000"
        assert records[0].dept_input == "QA-Dept"
        assert records[0].role_input == "R_DEV"
        assert records[0].user_gender == "1"
        assert records[0].status == "1"

    def test_value_only_import_suppresses_data_validation_extension_warning(self):
        rows = [["alice", "", "Alice", "", "", "QA", "", "0", "1"]]
        data = _xlsx_with_data_validation_extension(rows)

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            records = parse_import_excel(data, MIME_XLSX)

        assert records[0].user_name == "alice"
        assert not any(
            "Data Validation extension is not supported" in str(item.message)
            for item in caught
        )

    def test_parse_csv_basic(self):
        rows = [
            [
                "carol",
                "",
                "Carol",
                "carol@example.com",
                "13700137000",
                "QA-Dept",
                "",
                "0",
                "1",
            ],
        ]
        records = parse_import_excel(_csv_bytes(rows), MIME_CSV)
        assert len(records) == 1
        assert records[0].user_name == "carol"
        assert records[0].employee_no is None  # 空字符串规范化为 None。

    def test_parse_skips_empty_rows(self):
        rows = [
            ["alice", "", "Alice", "", "", "QA-Dept", "", "0", "1"],
            ["", "", "", "", "", "", "", "", ""],  # 全空行
            ["bob", "", "Bob", "", "", "QA-Dept2", "", "0", "1"],
        ]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert len(records) == 2
        assert {r.user_name for r in records} == {"alice", "bob"}


class TestOptionalFieldsDefaults:
    """可选字段缺省时的默认值（gender="0" / status="1"）。"""

    def test_gender_defaults_to_zero(self):
        rows = [["dave", "", "Dave", "", "", "QA-Dept", "", "", ""]]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].user_gender == "0"
        assert records[0].status == "1"  # 默认启用

    def test_employee_no_blank_normalized_to_none(self):
        """normalize_employee_no 将空值转为 None，避免 UNIQUE 冲突。"""
        rows = [["eve", "   ", "Eve", "", "", "QA-Dept", "", "0", "1"]]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].employee_no is None

    def test_role_input_blank_normalized_to_none(self):
        rows = [["frank", "", "Frank", "", "", "QA-Dept", "   ", "0", "1"]]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].role_input is None


class TestFieldValidationCollectsErrors:
    """失败行应汇总到一个 ImportErrorCollection。"""

    def test_missing_required_user_name_collected(self):
        rows = [
            ["", "", "NoName", "", "", "QA-Dept", "", "0", "1"],
        ]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert len(errors) == 1
        assert errors[0].row_num == 2
        assert errors[0].field == "user_name"
        assert errors[0].error_code == "AI_IMPORT_USERNAME_INVALID"

    def test_missing_required_dept_input_collected(self):
        rows = [
            ["alice", "", "Alice", "", "", "", "", "0", "1"],
        ]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(
            e.field == "dept_input" and e.error_code == "AI_IMPORT_DEPT_INPUT_REQUIRED"
            for e in errors
        )

    def test_user_name_too_short_collected(self):
        rows = [["a", "", "A", "", "", "QA-Dept", "", "0", "1"]]  # 1 字符 < 2
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(e.field == "user_name" for e in errors)

    def test_user_name_too_long_collected(self):
        rows = [["a" * 17, "", "Long", "", "", "QA-Dept", "", "0", "1"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(e.field == "user_name" for e in errors)

    def test_email_format_invalid_collected(self):
        """spec line 2630：邮箱格式错 → AI_IMPORT_EMAIL_INVALID。"""
        rows = [["alice", "", "Alice", "not-an-email", "", "QA-Dept", "", "0", "1"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert len(errors) == 1
        assert errors[0].field == "user_email"
        assert errors[0].error_code == "AI_IMPORT_EMAIL_INVALID"

    def test_phone_format_invalid_collected(self):
        rows = [["alice", "", "Alice", "", "12345", "QA-Dept", "", "0", "1"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert len(errors) == 1
        assert errors[0].field == "user_phone"
        assert errors[0].error_code == "AI_IMPORT_PHONE_INVALID"

    def test_phone_valid_mainland_format_passes(self):
        rows = [["alice", "", "Alice", "", "13800138000", "QA-Dept", "", "0", "1"]]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].user_phone == "13800138000"

    def test_gender_invalid_collected(self):
        rows = [["alice", "", "Alice", "", "", "QA-Dept", "", "9", "1"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(
            e.field == "user_gender" and e.error_code == "AI_IMPORT_GENDER_INVALID"
            for e in errors
        )

    def test_status_invalid_collected(self):
        rows = [["alice", "", "Alice", "", "", "QA-Dept", "", "0", "9"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(
            e.field == "status" and e.error_code == "AI_IMPORT_STATUS_INVALID"
            for e in errors
        )

    def test_collects_multiple_errors_across_rows(self):
        """一次收集所有错误，避免用户逐行修复和重传。"""
        rows = [
            ["", "", "NoName", "", "", "QA-Dept", "", "0", "1"],  # row 2 user_name 缺
            [
                "bob",
                "",
                "Bob",
                "bad-email",
                "",
                "QA-Dept",
                "",
                "0",
                "1",
            ],  # row 3 email 错
            [
                "carol",
                "",
                "Carol",
                "",
                "12345",
                "QA-Dept",
                "",
                "0",
                "1",
            ],  # row 4 phone 错
            ["dave", "", "Dave", "", "", "QA-Dept", "", "0", "1"],  # row 5 OK
        ]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        # 3 行各 1 个错误，row 5 OK
        assert len(errors) == 3
        row_nums = {e.row_num for e in errors}
        assert row_nums == {2, 3, 4}
        # row_num 升序
        assert [e.row_num for e in errors] == sorted(e.row_num for e in errors)

    def test_single_row_multiple_errors_collected(self):
        """同一行多个字段错 → 全部收集（user_name 缺 + email 错 + phone 错）。"""
        rows = [["", "", "X", "bad-email", "12345", "", "", "9", "9"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        fields = {e.field for e in errors}
        assert "user_name" in fields
        assert "user_email" in fields
        assert "user_phone" in fields
        assert "dept_input" in fields
        assert "user_gender" in fields
        assert "status" in fields


class TestImportErrorCollection:
    """验证 ImportErrorCollection 异常行为。"""

    def test_is_exception_subclass(self):
        exc = ImportErrorCollection(
            errors=[
                FailedRow(
                    row_num=2,
                    field="user_name",
                    value="",
                    reason="必填",
                    error_code="AI_IMPORT_USERNAME_INVALID",
                )
            ]
        )
        assert isinstance(exc, Exception)
        assert len(exc.errors) == 1
        assert "1 field errors" in str(exc)


# ========== 中文显示文本反查与 status 取值 ==========


class TestChineseLabelRoundTrip:
    """导出 Excel 的中文显示文本可以往返导入。

    场景：管理员导出 users.xlsx（status="启用"，user_gender="男"，dept_input="总公司/研发中心"），
    修改后重新导入 — status / user_gender 中文字面值必须被反查回字面值。
    """

    def test_status_chinese_label_round_trip(self):
        """status 中文标签 "启用" → "1"；"禁用" → "2"。"""
        rows = [
            ["alice", "", "Alice", "", "", "QA-Dept", "", "0", "启用"],
            ["bob", "", "Bob", "", "", "QA-Dept", "", "0", "禁用"],
        ]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].status == "1"
        assert records[1].status == "2"

    def test_gender_chinese_label_round_trip(self):
        """user_gender 中文标签 "未知"→"0"，"男"→"1"，"女"→"2"。"""
        rows = [
            ["alice", "", "Alice", "", "", "QA-Dept", "", "未知", "1"],
            ["bob", "", "Bob", "", "", "QA-Dept", "", "男", "1"],
            ["carol", "", "Carol", "", "", "QA-Dept", "", "女", "1"],
        ]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].user_gender == "0"
        assert records[1].user_gender == "1"
        assert records[2].user_gender == "2"

    def test_status_disabled_two_now_accepted(self):
        """_STATUS_VALUES 使用数据库真实取值 {"1", "2"}，
        DB 真实取值 status="2"（禁用）必须可导入。"""
        rows = [["alice", "", "Alice", "", "", "QA-Dept", "", "0", "2"]]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].status == "2"

    def test_status_zero_now_rejected(self):
        """status="0" 不属于数据库真实取值集合（"1", "2"），
        应抛 AI_IMPORT_STATUS_INVALID 让用户改 Excel（不能静默写错数据）。"""
        rows = [["alice", "", "Alice", "", "", "QA-Dept", "", "0", "0"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(
            e.field == "status" and e.error_code == "AI_IMPORT_STATUS_INVALID"
            for e in errors
        )

    def test_status_literal_still_works(self):
        """字面值 "1" / "2" 仍然合法（向后兼容非 round-trip 场景）。"""
        rows = [
            ["alice", "", "Alice", "", "", "QA-Dept", "", "0", "1"],
            ["bob", "", "Bob", "", "", "QA-Dept", "", "0", "2"],
        ]
        records = parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        assert records[0].status == "1"
        assert records[1].status == "2"

    def test_unknown_chinese_label_rejected(self):
        """非翻译字典中的中文字面值（如 "启用中"）应抛 AI_IMPORT_STATUS_INVALID。"""
        rows = [["alice", "", "Alice", "", "", "QA-Dept", "", "0", "启用中"]]
        with pytest.raises(ImportErrorCollection) as exc_info:
            parse_import_excel(_xlsx_bytes(rows), MIME_XLSX)
        errors = exc_info.value.errors
        assert any(
            e.field == "status" and e.error_code == "AI_IMPORT_STATUS_INVALID"
            for e in errors
        )
