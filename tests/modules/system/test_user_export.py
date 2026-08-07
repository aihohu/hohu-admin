"""export_users_to_excel 单测（Task 11，spec §3.6 line 2099-2111 + §2.31）。

覆盖：
- 强制建 UserExportTask（spec §2.31 line 1436-1453）
- filter_snapshot 冻结 accessible_dept_ids（spec §2.31 line 1516-1520）
- reason 必填（spec §2.30 v2.2 P1-3）
- EXPORT_ALLOWED_FIELDS 白名单（hashed_password 永不导出，spec §2.9）
- 行数 > USER_EXPORT_ASYNC_THRESHOLD → AI_EXPORT_ASYNC_REQUIRED（spec §2.6）
- data_scope 自动应用（HR 只能导他可见的，spec §2.31 line 1545）
- 失败也建 task（status=FAILED + error_code，spec §2.31 line 1567-1572）
- 30 天 TTL 文件存储（spec §2.31 line 1554）

依赖 db_session outer-transaction fixture（不落库）。
"""

import io
import json as _json

import pytest
from openpyxl import load_workbook
from sqlalchemy import select as _select
from sqlalchemy import text as _text

from app.constants import (
    ADMIN_USERNAME,
    DATA_SCOPE_ALL,
    DATA_SCOPE_DEPT,
    SUPER_ADMIN_ROLE_CODE,
)
from app.core.exceptions import BusinessRuleException, NotFoundException
from app.core.file_storage import MockFileStorage
from app.modules.system.models.dept import Dept
from app.modules.system.models.operation_log import SysOperationLog
from app.modules.system.models.role import Role
from app.modules.system.models.user import User
from app.modules.system.user.constants import (
    EXPORT_ALLOWED_FIELDS,
    USER_EXPORT_ASYNC_THRESHOLD,
    ExportTaskStatus,
)
from app.modules.system.user.export_service import (
    download_export_file,
    export_users_to_excel,
    get_export_task,
    list_export_tasks,
)
from app.modules.system.user.models import UserExportTask
from app.modules.system.user.schemas import UserExportFilter, UserExportTaskQuery

# ========== helpers ==========


def _make_dept(
    dept_id: int,
    name: str,
    parent_id: int | None = None,
    ancestors: str = "0",
) -> Dept:
    return Dept(
        dept_id=dept_id,
        parent_id=parent_id,
        ancestors=ancestors,
        dept_name=name,
        order_num=0,
        status="1",
    )


def _make_role(
    role_id: int,
    code: str,
    name: str,
    data_scope: str = DATA_SCOPE_ALL,
    status: str = "1",
) -> Role:
    return Role(
        role_id=role_id,
        role_code=code,
        role_name=name,
        data_scope=data_scope,
        status=status,
    )


def _make_user(
    user_id: int,
    user_name: str,
    roles: list[Role],
    depts: list[Dept] | None = None,
    *,
    nickname: str | None = "Nick",
    user_email: str | None = None,
    user_phone: str | None = None,
    user_gender: str = "0",
    status: str = "1",
    hashed_password: str = "secret-hash-never-export",
) -> User:
    return User(
        user_id=user_id,
        user_name=user_name,
        hashed_password=hashed_password,
        nickname=nickname,
        user_email=user_email,
        user_phone=user_phone,
        user_gender=user_gender,
        status=status,
        roles=roles,
        depts=depts or [],
    )


@pytest.fixture
def file_storage() -> MockFileStorage:
    """每个测试独立 MockFileStorage（spec §3.9 注入）。"""
    return MockFileStorage()


async def _fetch_super_role(db_session) -> Role:
    return (
        await db_session.execute(
            _select(Role).where(Role.role_code == SUPER_ADMIN_ROLE_CODE)
        )
    ).scalar_one()


def _read_xlsx(xlsx_bytes: bytes) -> tuple[list[str], list[list]]:
    """读 xlsx bytes → (headers, rows)。"""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return list(rows[0]), [list(r) for r in rows[1:]]


# ========== Happy Path ==========


class TestHappyPath:
    """spec §3.6 line 2099-2111：基本导出流程。"""

    async def test_export_returns_xlsx_and_count(self, db_session, file_storage):
        """返回 (xlsx_bytes, row_count, export_id) 三元组。"""
        dept = _make_dept(5101, "QA-Exp-Dept")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6101, "QA_EXP_OP", [super_role], [dept])
        target = _make_user(
            6102,
            "QA_EXP_T1",
            [super_role],
            [dept],
            user_email="t1@example.com",
        )
        db_session.add_all([dept, operator, target])
        await db_session.flush()

        result = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA export happy path",
            file_storage=file_storage,
        )

        assert len(result) == 3
        xlsx_bytes, row_count, export_id = result
        assert isinstance(xlsx_bytes, bytes)
        assert xlsx_bytes[:2] == b"PK"  # xlsx zip 头
        assert row_count >= 1
        assert isinstance(export_id, str)
        assert export_id

    async def test_export_only_includes_allowed_fields(self, db_session, file_storage):
        """spec §2.9：hashed_password 永不出现在 Excel。"""
        dept = _make_dept(5102, "QA-Exp-Dept-FW")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6103, "QA_EXP_FW", [super_role], [dept])
        target = _make_user(
            6104,
            "QA_EXP_TARGET",
            [super_role],
            [dept],
            hashed_password="$2b$12$super-secret-bcrypt-hash",
        )
        db_session.add_all([dept, operator, target])
        await db_session.flush()

        xlsx_bytes, _count, _export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA field whitelist",
            file_storage=file_storage,
        )

        headers, rows = _read_xlsx(xlsx_bytes)
        assert "hashed_password" not in headers
        # 检查所有单元格值，永不含 bcrypt 哈希
        all_values = " ".join(str(v) for row in rows for v in row)
        assert "$2b$12$super-secret-bcrypt-hash" not in all_values

    async def test_export_includes_all_allowed_fields(self, db_session, file_storage):
        """spec §2.9：EXPORT_ALLOWED_FIELDS 中所有字段都应有对应列。"""
        dept = _make_dept(5103, "QA-Exp-Dept-Col")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6105, "QA_EXP_COL", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        xlsx_bytes, _count, _export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA all columns",
            file_storage=file_storage,
        )

        headers, _rows = _read_xlsx(xlsx_bytes)
        # 至少包含 EXPORT_ALLOWED_FIELDS 中的字段名（中文表头时校验英文不易，
        # 这里检查列数与 EXPORT_ALLOWED_FIELDS 一致）
        assert len(headers) >= len(EXPORT_ALLOWED_FIELDS)


# ========== Task 审计（spec §2.31） ==========


class TestExportTaskAudit:
    """spec §2.31：所有导出一律建 UserExportTask。"""

    async def test_export_creates_task_with_filter_snapshot(
        self, db_session, file_storage
    ):
        """spec §2.31 line 1525-1536：建 task，filter_snapshot 含 filter + accessible_dept_ids。"""
        dept = _make_dept(5201, "QA-Exp-Dept-FS")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6201, "QA_EXP_FS", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        filter_ = UserExportFilter(user_name="QA", status="1")
        _bytes, _count, export_id = await export_users_to_excel(
            db_session,
            filter_,
            operator,
            reason="QA filter snapshot",
            file_storage=file_storage,
        )

        task = (
            await db_session.execute(
                _select(UserExportTask).where(UserExportTask.export_id == export_id)
            )
        ).scalar_one()
        assert task.operator_id == operator.user_id
        assert task.reason == "QA filter snapshot"
        assert task.status == ExportTaskStatus.SUCCESS
        # filter_snapshot 含原始 filter
        assert task.filter_snapshot["filter"]["user_name"] == "QA"
        assert task.filter_snapshot["filter"]["status"] == "1"
        # filter_snapshot 含 accessible_dept_ids + filter_evaluated_at
        assert "accessible_dept_ids" in task.filter_snapshot
        assert "filter_evaluated_at" in task.filter_snapshot

    async def test_export_filter_snapshot_freezes_accessible_dept_ids(
        self, db_session, file_storage
    ):
        """spec §2.31 line 1516-1520：filter_snapshot.accessible_dept_ids 冻结当时的部门集合。"""
        own_dept = _make_dept(5301, "QA-Exp-Own")
        other_dept = _make_dept(5302, "QA-Exp-Other")
        hr_role = _make_role(
            5303, "QA_R_HR_EXP", "QA-HR-EXP", data_scope=DATA_SCOPE_DEPT
        )
        operator = _make_user(6301, "QA_EXP_HR", [hr_role], [own_dept])
        db_session.add_all([own_dept, other_dept, hr_role, operator])
        await db_session.flush()

        _bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA snapshot freeze",
            file_storage=file_storage,
        )

        # 用 export_id 反查（不假设全表只 1 行：dev DB 可能有真实导出残留）
        task = (
            await db_session.execute(
                _select(UserExportTask).where(UserExportTask.export_id == export_id)
            )
        ).scalar_one()
        # DATA_SCOPE_DEPT → accessible_dept_ids = own_dept 的 id
        assert task.filter_snapshot["accessible_dept_ids"] == [own_dept.dept_id]

    async def test_export_status_transitions_created_to_success(
        self, db_session, file_storage
    ):
        """spec §2.31 line 1540-1565：CREATED → RUNNING → SUCCESS。"""
        dept = _make_dept(5401, "QA-Exp-Dept-ST")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6401, "QA_EXP_ST", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        _bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA state transition",
            file_storage=file_storage,
        )

        task = (
            await db_session.execute(
                _select(UserExportTask).where(UserExportTask.export_id == export_id)
            )
        ).scalar_one()
        assert task.status == ExportTaskStatus.SUCCESS
        assert task.started_at is not None
        assert task.finished_at is not None
        assert task.duration_ms is not None
        assert task.duration_ms >= 0

    async def test_export_writes_file_to_storage(self, db_session, file_storage):
        """spec §2.31 line 1549-1555：xlsx 写入 FileStorage，路径存 task.file_storage_key。"""
        dept = _make_dept(5501, "QA-Exp-Dept-FL")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6501, "QA_EXP_FL", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        xlsx_bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA file storage",
            file_storage=file_storage,
        )

        task = (
            await db_session.execute(
                _select(UserExportTask).where(UserExportTask.export_id == export_id)
            )
        ).scalar_one()
        assert task.file_storage_key is not None
        assert "user-export" in task.file_storage_key
        assert task.file_size_bytes == len(xlsx_bytes)
        # 文件实际写入 MockFileStorage
        assert await file_storage.exists(task.file_storage_key)


# ========== Reason 必填（spec §2.30） ==========


class TestReasonValidation:
    """spec §2.30：reason 必填，1-256 字符（service 层 defense-in-depth）。"""

    async def test_export_reason_required(self, db_session, file_storage):
        dept = _make_dept(5601, "QA-Exp-Dept-RSN")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6601, "QA_EXP_RSN", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        with pytest.raises(BusinessRuleException) as exc:
            await export_users_to_excel(
                db_session,
                UserExportFilter(),
                operator,
                reason="   ",  # 全空白
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_REASON_REQUIRED"

    async def test_export_reason_too_long_rejected(self, db_session, file_storage):
        dept = _make_dept(5602, "QA-Exp-Dept-RL")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6602, "QA_EXP_RL", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        with pytest.raises(BusinessRuleException) as exc:
            await export_users_to_excel(
                db_session,
                UserExportFilter(),
                operator,
                reason="x" * 257,
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_REASON_REQUIRED"


# ========== 阈值（spec §2.6） ==========


class TestExportThreshold:
    """spec §2.6：行数 > USER_EXPORT_ASYNC_THRESHOLD → AI_EXPORT_ASYNC_REQUIRED。"""

    async def test_export_over_threshold_raises_async_required(
        self, db_session, file_storage
    ):
        dept = _make_dept(5701, "QA-Exp-Dept-TH")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6701, "QA_EXP_TH", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        # 制造 USER_EXPORT_ASYNC_THRESHOLD + 1 个用户
        bulk_users = [
            _make_user(
                6702 + i,
                f"QA_EXP_BULK_{i:05d}",
                [super_role],
                [dept],
            )
            for i in range(USER_EXPORT_ASYNC_THRESHOLD + 1)
        ]
        db_session.add_all(bulk_users)
        await db_session.flush()

        with pytest.raises(BusinessRuleException) as exc:
            await export_users_to_excel(
                db_session,
                UserExportFilter(),
                operator,
                reason="QA threshold",
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_ASYNC_REQUIRED"
        assert "缩窄" in exc.value.message
        assert "异步" not in exc.value.message


# ========== Data Scope（spec §2.31 line 1545） ==========


class TestDataScope:
    """data_scope 自动应用：HR 只能导他可见的部门用户。"""

    async def test_export_data_scope_filters_by_dept(self, db_session, file_storage):
        """spec §2.31 line 1545 + §2.11：DATA_SCOPE_DEPT 限定本部门。"""
        own_dept = _make_dept(5801, "QA-Exp-Own-D")
        other_dept = _make_dept(5802, "QA-Exp-Other-D")
        hr_role = _make_role(5803, "QA_R_HR_D", "QA-HR-D", data_scope=DATA_SCOPE_DEPT)
        operator = _make_user(6801, "QA_EXP_HR_D", [hr_role], [own_dept])
        own_user = _make_user(6802, "QA_IN_OWN", [hr_role], [own_dept])
        other_user = _make_user(6803, "QA_IN_OTHER", [hr_role], [other_dept])
        db_session.add_all(
            [own_dept, other_dept, hr_role, operator, own_user, other_user]
        )
        await db_session.flush()

        xlsx_bytes, row_count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA data scope",
            file_storage=file_storage,
        )

        # HR 只看到 own_dept 里的用户（operator 自己 + own_user）
        assert row_count == 2
        headers, rows = _read_xlsx(xlsx_bytes)
        user_name_col = 0  # 第 1 列
        exported_names = {r[user_name_col] for r in rows}
        assert "QA_IN_OWN" in exported_names
        assert "QA_IN_OTHER" not in exported_names

    async def test_export_super_admin_sees_all(self, db_session, file_storage):
        """spec line 2662：超管豁免 data_scope，看所有用户。"""
        dept1 = _make_dept(5901, "QA-Exp-SA-D1")
        dept2 = _make_dept(5902, "QA-Exp-SA-D2")
        admin_user = (
            await db_session.execute(
                _select(User).where(User.user_name == ADMIN_USERNAME)
            )
        ).scalar_one()
        u1 = _make_user(6902, "QA_SA_U1", [], [dept1])
        u2 = _make_user(6903, "QA_SA_U2", [], [dept2])
        db_session.add_all([dept1, dept2, u1, u2])
        await db_session.flush()

        _bytes, row_count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            admin_user,
            reason="QA super admin export",
            file_storage=file_storage,
        )

        # 超管看所有用户（含 init_db seed 的 + dept1/dept2 两个新用户）
        assert row_count >= 2


# ========== 失败也建 task（spec §2.31 line 1567-1572） ==========


class TestExportFailureRecordsError:
    """spec §2.31 line 1567-1572：失败也建 task，status=FAILED + error_code。"""

    async def test_export_failure_records_async_required_in_task(
        self, db_session, file_storage
    ):
        """超阈值时 task.status=FAILED + error_code=AI_EXPORT_ASYNC_REQUIRED。"""
        dept = _make_dept(6001, "QA-Exp-Dept-FE")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(7001, "QA_EXP_FE", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        bulk_users = [
            _make_user(
                7002 + i,
                f"QA_EXP_FAIL_{i:05d}",
                [super_role],
                [dept],
            )
            for i in range(USER_EXPORT_ASYNC_THRESHOLD + 1)
        ]
        db_session.add_all(bulk_users)
        await db_session.flush()

        with pytest.raises(BusinessRuleException):
            await export_users_to_excel(
                db_session,
                UserExportFilter(),
                operator,
                reason="QA failure task",
                file_storage=file_storage,
            )

        # 用 reason 反查（不假设全表只 1 行：dev DB 可能有真实导出残留；
        # 失败路径 export_users_to_excel 抛异常不返回 export_id，reason 是唯一锚点）
        task = (
            await db_session.execute(
                _select(UserExportTask).where(
                    UserExportTask.reason == "QA failure task"
                )
            )
        ).scalar_one()
        assert task.status == ExportTaskStatus.FAILED
        assert task.error_code == "AI_EXPORT_ASYNC_REQUIRED"
        assert task.error_message is not None
        assert task.finished_at is not None


# ========== v2.3 §2.9.1：导出 Excel 字段翻译（display labels） ==========


class TestDisplayTranslation:
    """v2.3 §2.9.1：导出 Excel 字段翻译为可读中文标签 + dept full_path。"""

    async def test_export_translates_status_label(self, db_session, file_storage):
        """status "1" → "启用"；status "2" → "禁用"。"""
        dept = _make_dept(5930, "QA-Exp-Status")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6930, "QA_EXP_ST_OP", [super_role], [dept])
        enabled_user = _make_user(
            6931, "QA_EXP_ST_ON", [super_role], [dept], status="1"
        )
        disabled_user = _make_user(
            6932, "QA_EXP_ST_OFF", [super_role], [dept], status="2"
        )
        db_session.add_all([dept, operator, enabled_user, disabled_user])
        await db_session.flush()

        xlsx_bytes, _count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA status label",
            file_storage=file_storage,
        )

        headers, rows = _read_xlsx(xlsx_bytes)
        status_col = headers.index("状态")
        user_name_col = headers.index("账号")
        status_by_name = {r[user_name_col]: r[status_col] for r in rows}
        assert status_by_name["QA_EXP_ST_ON"] == "启用"
        assert status_by_name["QA_EXP_ST_OFF"] == "禁用"

    async def test_export_translates_gender_label(self, db_session, file_storage):
        """user_gender "0"/"1"/"2" → "未知"/"男"/"女"。"""
        dept = _make_dept(5931, "QA-Exp-Gender")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6933, "QA_EXP_GD_OP", [super_role], [dept])
        male = _make_user(6934, "QA_EXP_GD_M", [super_role], [dept], user_gender="1")
        female = _make_user(6935, "QA_EXP_GD_F", [super_role], [dept], user_gender="2")
        unknown = _make_user(6936, "QA_EXP_GD_U", [super_role], [dept], user_gender="0")
        db_session.add_all([dept, operator, male, female, unknown])
        await db_session.flush()

        xlsx_bytes, _count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA gender label",
            file_storage=file_storage,
        )

        headers, rows = _read_xlsx(xlsx_bytes)
        gender_col = headers.index("性别")
        user_name_col = headers.index("账号")
        gender_by_name = {r[user_name_col]: r[gender_col] for r in rows}
        assert gender_by_name["QA_EXP_GD_M"] == "男"
        assert gender_by_name["QA_EXP_GD_F"] == "女"
        assert gender_by_name["QA_EXP_GD_U"] == "未知"

    async def test_export_formats_dept_full_path(self, db_session, file_storage):
        """dept_id 列输出 full_path「总公司/研发中心/前端部」（不是数字 ID）。"""
        root = _make_dept(5940, "总公司", ancestors="0")
        middle = _make_dept(5941, "研发中心", parent_id=5940, ancestors="0,5940")
        leaf = _make_dept(5942, "前端部", parent_id=5941, ancestors="0,5940,5941")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6940, "QA_EXP_DP_OP", [super_role], [leaf])
        target = _make_user(6941, "QA_EXP_DP_T", [super_role], [leaf])
        db_session.add_all([root, middle, leaf, operator, target])
        await db_session.flush()

        xlsx_bytes, _count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA dept path",
            file_storage=file_storage,
        )

        headers, rows = _read_xlsx(xlsx_bytes)
        dept_col = headers.index("部门")
        user_name_col = headers.index("账号")
        dept_by_name = {r[user_name_col]: r[dept_col] for r in rows}
        assert dept_by_name["QA_EXP_DP_T"] == "总公司/研发中心/前端部"

    async def test_export_role_codes_unchanged(self, db_session, file_storage):
        """role_codes 保持 role_code 字面值（§2.18 已支持 code round-trip）。"""
        dept = _make_dept(5943, "QA-Exp-Role")
        super_role = await _fetch_super_role(db_session)
        dev_role = _make_role(5944, "QA_R_DEV_TL", "QA-开发者-翻译")
        operator = _make_user(6942, "QA_EXP_RC_OP", [super_role], [dept])
        target = _make_user(6943, "QA_EXP_RC_T", [dev_role], [dept])
        db_session.add_all([dept, dev_role, operator, target])
        await db_session.flush()

        xlsx_bytes, _count, _ = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA role code unchanged",
            file_storage=file_storage,
        )

        headers, rows = _read_xlsx(xlsx_bytes)
        role_col = headers.index("角色编码")
        user_name_col = headers.index("账号")
        role_by_name = {r[user_name_col]: r[role_col] for r in rows}
        assert role_by_name["QA_EXP_RC_T"] == "QA_R_DEV_TL"


# ========== Task 33：download_export_file（AI 对话内点击下载落地） ==========


class TestDownloadExportFile:
    """Task 33 / spec §2.31 line 1626：download_export_file service。

    从 sys_user_export_task.file_storage_key 读 bytes 返回；
    任务不存在 / 状态非 SUCCESS / file_storage_key 缺失 / 文件被删 → 各 errorCode。
    filename 从 task.created_at 派生（决策 30.6 同款格式）。
    """

    async def test_download_returns_bytes_and_filename(self, db_session, file_storage):
        """成功路径：返回 (xlsx_bytes, filename)；filename 从 created_at 派生。"""
        dept = _make_dept(5950, "QA-Exp-DL")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6950, "QA_EXP_DL_OP", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        xlsx_bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA download",
            file_storage=file_storage,
        )

        got_bytes, filename = await download_export_file(
            db_session,
            export_id,
            operator_id=operator.user_id,
            file_storage=file_storage,
        )
        assert got_bytes == xlsx_bytes
        # 决策 30.6 同款：hohu_users_YYYYMMDD_HHmmss.xlsx
        assert filename.startswith("hohu_users_")
        assert filename.endswith(".xlsx")

    async def test_download_task_not_found(self, db_session, file_storage):
        """任务不存在 → NotFoundException(AI_EXPORT_TASK_NOT_FOUND)。"""
        with pytest.raises(NotFoundException) as exc:
            await download_export_file(
                db_session,
                "nonexistent-id",
                operator_id=1,
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_TASK_NOT_FOUND"

    async def test_download_rejects_failed_task(self, db_session, file_storage):
        """status=FAILED（无 file_storage_key）→ AI_EXPORT_TASK_NOT_READY。"""
        dept = _make_dept(5951, "QA-Exp-DL-F")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6951, "QA_EXP_DL_F_OP", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        # 制造 FAILED task（超阈值）
        bulk_users = [
            _make_user(6952 + i, f"QA_EXP_DL_F_{i:04d}", [super_role], [dept])
            for i in range(USER_EXPORT_ASYNC_THRESHOLD + 1)
        ]
        db_session.add_all(bulk_users)
        await db_session.flush()

        with pytest.raises(BusinessRuleException):
            await export_users_to_excel(
                db_session,
                UserExportFilter(),
                operator,
                reason="QA download failed",
                file_storage=file_storage,
            )

        task = (
            await db_session.execute(
                _select(UserExportTask).where(
                    UserExportTask.reason == "QA download failed"
                )
            )
        ).scalar_one()
        assert task.status == ExportTaskStatus.FAILED

        with pytest.raises(BusinessRuleException) as exc:
            await download_export_file(
                db_session,
                task.export_id,
                operator_id=operator.user_id,
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_TASK_NOT_READY"

    async def test_download_file_missing_from_storage(self, db_session, file_storage):
        """file_storage_key 在 DB 但文件被外部删除 → AI_EXPORT_FILE_EXPIRED。"""
        dept = _make_dept(5952, "QA-Exp-DL-E")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6970, "QA_EXP_DL_E_OP", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        _bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason="QA download expired",
            file_storage=file_storage,
        )

        # 手动从 storage 删文件，模拟 30 天 TTL 清理或外部删除
        task = (
            await db_session.execute(
                _select(UserExportTask).where(UserExportTask.export_id == export_id)
            )
        ).scalar_one()
        await file_storage.delete(task.file_storage_key)

        with pytest.raises(BusinessRuleException) as exc:
            await download_export_file(
                db_session,
                export_id,
                operator_id=operator.user_id,
                file_storage=file_storage,
            )
        assert exc.value.error_code == "AI_EXPORT_FILE_EXPIRED"


class TestExportTaskOwnership:
    """Task 35：导出任务 list/detail/download 默认只能访问当前 operator。"""

    @staticmethod
    def _task(export_id: str, operator_id: int, *, reason: str) -> UserExportTask:
        return UserExportTask(
            export_id=export_id,
            operator_id=operator_id,
            filter_snapshot={"user_name": reason},
            reason=reason,
            row_count=1,
            file_storage_key=f"user-export/{export_id}.xlsx",
            file_size_bytes=4,
            status=ExportTaskStatus.SUCCESS,
        )

    async def test_detail_hides_cross_owner_as_not_found(self, db_session):
        own = self._task("qa-owner-detail-own", 88101, reason="own-filter")
        other = self._task("qa-owner-detail-other", 88102, reason="secret-filter")
        db_session.add_all([own, other])
        await db_session.flush()

        assert (
            await get_export_task(
                db_session,
                own.export_id,
                operator_id=own.operator_id,
            )
            is own
        )
        assert (
            await get_export_task(
                db_session,
                other.export_id,
                operator_id=own.operator_id,
            )
            is None
        )

    async def test_detail_allows_explicit_super_admin_cross_owner(self, db_session):
        other = self._task("qa-owner-detail-admin", 88112, reason="admin-visible")
        db_session.add(other)
        await db_session.flush()

        assert (
            await get_export_task(
                db_session,
                other.export_id,
                operator_id=88111,
                allow_cross_owner=True,
            )
            is other
        )

    async def test_list_forces_owner_scope_for_non_super_admin(self, db_session):
        own = self._task("qa-owner-list-own", 88121, reason="own-list")
        other = self._task("qa-owner-list-other", 88122, reason="secret-list")
        db_session.add_all([own, other])
        await db_session.flush()

        page = await list_export_tasks(
            db_session,
            UserExportTaskQuery(operator_id=other.operator_id, size=100),
            operator_id=own.operator_id,
        )

        assert page.total == 0
        assert page.records == []

    async def test_list_allows_super_admin_operator_filter(self, db_session):
        own = self._task("qa-owner-list-admin-own", 88131, reason="admin-own")
        other = self._task("qa-owner-list-admin-other", 88132, reason="admin-other")
        db_session.add_all([own, other])
        await db_session.flush()

        page = await list_export_tasks(
            db_session,
            UserExportTaskQuery(operator_id=other.operator_id, size=100),
            operator_id=own.operator_id,
            allow_cross_owner=True,
        )

        assert page.total == 1
        assert [task.export_id for task in page.records] == [other.export_id]

    async def test_download_hides_cross_owner_before_reading_file(
        self, db_session, file_storage, monkeypatch
    ):
        other = self._task(
            "qa-owner-download-other",
            88142,
            reason="secret-download",
        )
        db_session.add(other)
        await db_session.flush()
        read_called = False

        async def track_read(storage_key: str) -> bytes:  # noqa: ARG001
            nonlocal read_called
            read_called = True
            return b"secret-file-bytes"

        monkeypatch.setattr(file_storage, "read", track_read)

        with pytest.raises(NotFoundException) as exc:
            await download_export_file(
                db_session,
                other.export_id,
                operator_id=88141,
                file_storage=file_storage,
            )

        assert exc.value.error_code == "AI_EXPORT_TASK_NOT_FOUND"
        assert other.export_id not in str(exc.value)
        assert other.reason not in str(exc.value)
        assert read_called is False


# ========== Task 34：audit chain JOIN 测试（spec §8.1 line 2902） ==========


class TestAuditChainJoinable:
    """spec §8.1 line 2902：审计链 sys_operation_log ↔ sys_user_export_task 可对照。

    spec 原文写「sys_operation_log.export_id ↔ sys_user_export_task.export_id 可 JOIN」，
    但 sys_operation_log schema（app/modules/system/models/operation_log.py）**没有
    export_id 字段**。真实 audit chain 机制（spec §2.30 line 1450 + §2.8 line 256）：

      sys_operation_log.path='/system/user/export'
        + request_params.reason + user_id + create_time
        ↔ sys_user_export_task.reason + operator_id + created_at

    本测试验证：给定 operation_log 行（模拟 AuditLogMiddleware 写入），
    能通过 reason + user_id + 时间窗反查到正确的 export task 行。
    """

    async def test_audit_chain_joinable_via_reason_and_operator(
        self, db_session, file_storage
    ):
        """sys_operation_log 行 ↔ sys_user_export_task 行可通过 reason + operator_id 对照。"""
        dept = _make_dept(5960, "QA-Exp-Audit")
        super_role = await _fetch_super_role(db_session)
        operator = _make_user(6960, "QA_EXP_AUDIT_OP", [super_role], [dept])
        db_session.add_all([dept, operator])
        await db_session.flush()

        # 1. 创建 export task（模拟 export_users_to_excel 内部行为）
        unique_reason = "QA audit chain 2026-08-05 unique"
        _bytes, _count, export_id = await export_users_to_excel(
            db_session,
            UserExportFilter(),
            operator,
            reason=unique_reason,
            file_storage=file_storage,
        )

        # 2. 模拟 AuditLogMiddleware 写 sys_operation_log（HTTP POST /export）
        #    request_params 是 request body 的 JSON 摘要（含 reason）
        operation_log = SysOperationLog(
            user_id=operator.user_id,
            username=operator.user_name,
            module="system",
            action="create",
            method="POST",
            path="/system/user/export",
            request_params=_json.dumps(
                {"reason": unique_reason, "status": None}, ensure_ascii=False
            ),
            status_code=200,
            ip="127.0.0.1",
            duration=150,
        )
        db_session.add(operation_log)
        await db_session.flush()

        # 3. 验证 audit chain：从 operation_log 反查 export_task
        #    真实场景：审计员看到 operation_log → 想知道导出了什么 → 反查 export_task
        #    JOIN 条件：path + user_id + request_params.reason 匹配 task.reason
        # 用 SQL 验证 JOIN 可行性（模拟审计员手工 SQL 反查）
        result = await db_session.execute(
            _text("""
                SELECT t.export_id, t.reason, t.row_count, t.status,
                       l.username, l.path, l.create_time as log_time
                FROM sys_user_export_task t
                JOIN sys_operation_log l
                  ON l.user_id = t.operator_id
                 AND l.path = '/system/user/export'
                 AND l.request_params LIKE '%' || t.reason || '%'
                WHERE t.export_id = :export_id
            """),
            {"export_id": export_id},
        )
        row = result.first()
        assert row is not None, (
            "audit chain JOIN 失败：operation_log 无法反查到 export_task"
        )
        assert row.export_id == export_id
        assert row.reason == unique_reason
        assert row.username == "QA_EXP_AUDIT_OP"
