"""``GET /system/user/import/template`` HTTP 契约测试。

只验证 HTTP 契约层（路由 / streaming 响应 / Content-Disposition / 错误码）+
xlsx 内部结构（4 sheet / 列顺序 / DataValidation / 字典实时数据）。
service 层完整业务逻辑（部门字典实时查询 / 角色字典实时查询）走真实 service，
不 mock — 反而验证 endpoint 整合 service 的契约面。

覆盖：
- 401 未登录 / 无效 JWT（auth gating）
- 200 + Content-Type=xlsx + Content-Disposition: attachment; filename=user_import_template.xlsx
- 4 sheet：数据 / 说明 / 部门字典 / 角色字典
- 「数据」sheet 列顺序固定（user_name / nickname / user_email / user_phone /
  dept_input / role_input / user_gender / status）
- 「数据」sheet 包含两行示例，避免只有空表头
- 「数据」sheet 部门列（E）+ 角色列（F）有 DataValidation 下拉
- 「部门字典」sheet 实时查 sys_dept（含 full_path 列）+ 生成时间标注
- 「角色字典」sheet 实时查 sys_role + 生成时间标注
"""

import re
from datetime import UTC, datetime, timedelta
from io import BytesIO

import pytest
from httpx import ASGITransport, AsyncClient
from jose import jwt
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.config import settings
from app.main import app
from app.modules.system.models.dept import Dept
from app.modules.system.models.role import Role
from app.modules.system.models.user import User

#: xlsx MIME 类型。
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: xlsx 文件 magic bytes（PK zip header）
_XLSX_MAGIC = b"\x50\x4b\x03\x04"

#: 数据 sheet 的固定列顺序。
_EXPECTED_DATA_COLUMNS: tuple[str, ...] = (
    "user_name",
    "nickname",
    "user_email",
    "user_phone",
    "dept_input",
    "role_input",
    "user_gender",
    "status",
)

#: 四个 sheet 的名称。
_EXPECTED_SHEET_NAMES: tuple[str, ...] = (
    "数据",
    "说明",
    "部门字典",
    "角色字典",
)


# ========== Fixtures ==========


@pytest.fixture
async def client(db_session):  # noqa: ARG001 (db_session resets redis)
    """ASGI test client。

    依赖 ``db_session`` 触发 ``tests/modules/system/conftest.py`` 的
    ``_reset_redis_client()``，刷新 audit_middleware / auth.service 的 redis_client
    绑定到当前 loop（与 test_user_export_api.py / test_user_import_api.py 同款）。
    """
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as ac:
        yield ac


@pytest.fixture
async def admin_token(db_session) -> str:
    """admin 用户 JWT（admin 绕过 system:user:import 检查，决策 12.8）。"""
    user = (
        await db_session.execute(select(User).where(User.user_name == "admin"))
    ).scalar_one()
    exp = datetime.now(UTC) + timedelta(hours=1)
    payload = {
        "exp": exp,
        "sub": str(user.user_id),
        "tid": str(user.tenant_id),
        "type": "access",
        "user_id": user.user_id,
        "user_name": user.user_name,
    }
    return jwt.encode(payload, settings.SECRET_KEY, algorithm=settings.ALGORITHM)


# ========== Helpers ==========


def _load_xlsx(content: bytes):
    """加载 xlsx bytes → Workbook（断言 body 是合法 xlsx）。"""
    assert content.startswith(_XLSX_MAGIC), "响应 body 不是合法 xlsx（PK magic 缺失）"
    return load_workbook(BytesIO(content), data_only=False)


async def _seed_dept_and_role(db_session) -> tuple[Dept, Role]:
    """插入 1 个 dept + 1 个 role 用于字典 sheet 实时查询断言。

    init_db seed 的菜单 / 用户已存在，但 sys_dept / sys_role 不一定有适合本测试
    的稳定数据；这里显式建可读名字避免依赖 init_db seed 细节。
    """
    dept = Dept(
        dept_name="模板测试部门",
        parent_id=None,
        ancestors="0",
        status="1",
        order_num=999,
    )
    db_session.add(dept)
    role = Role(
        role_name="模板测试角色",
        role_code="R_TEMPLATE_TEST",
        status="1",
        data_scope="1",
    )
    db_session.add(role)
    await db_session.flush()
    return dept, role


# ========== Auth ==========


class TestDownloadTemplateAuth:
    """验证 system:user:import 权限。"""

    async def test_no_token_returns_401(self, client):
        response = await client.get("/system/user/import/template")
        assert response.status_code == 401

    async def test_invalid_token_returns_401(self, client):
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": "Bearer invalid.jwt.token"},
        )
        assert response.status_code == 401


# ========== 响应契约 ==========


class TestDownloadTemplateResponse:
    """响应返回 xlsx 文件和 Content-Disposition。"""

    async def test_returns_xlsx_with_content_disposition(self, client, admin_token):
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )

        assert response.status_code == 200, response.text
        # Content-Type 是 xlsx MIME
        assert response.headers["content-type"] == XLSX_MIME
        # Content-Disposition: attachment; filename=user_import_template.xlsx
        cd = response.headers.get("content-disposition", "")
        assert "attachment" in cd, f"Content-Disposition 缺 attachment: {cd}"
        match = re.search(r"filename=user_import_template\.xlsx", cd)
        assert match is not None, f"filename 不符合 spec: {cd}"
        # body 是合法 xlsx
        assert response.content.startswith(_XLSX_MAGIC)


# ========== Sheet 结构 ==========


class TestTemplateSheets:
    """模板包含四个 sheet；数据和说明固定，
    「部门字典」/「角色字典」实时查 DB 填充。
    """

    async def test_has_four_sheets_in_expected_order(self, client, admin_token):
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        sheet_names = wb.sheetnames
        assert sheet_names == list(_EXPECTED_SHEET_NAMES), (
            f"sheet 名称 / 顺序错: {sheet_names}"
        )

    async def test_data_sheet_has_correct_columns(self, client, admin_token):
        """数据 sheet 第一行使用固定英文列名。

        英文表头与 import_parser.EXCEL_HEADERS 对齐（决策 11.11：导入模板用英文），
        parser 按表头名匹配列索引。
        """
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        ws = wb["数据"]
        header = [cell.value for cell in ws[1]]
        # employee_no 是可选同步字段，不进入主模板的固定八列表头。
        assert header == list(_EXPECTED_DATA_COLUMNS), f"「数据」sheet 表头错: {header}"

    async def test_data_sheet_has_two_example_rows(self, client, admin_token):
        """模板提供两行示例，帮助用户理解必填列和格式。"""
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        ws = wb["数据"]
        # 第 2、3 行是示例（user_name 必填，至少要有值）
        row2_user_name = ws.cell(row=2, column=1).value
        row3_user_name = ws.cell(row=3, column=1).value
        assert row2_user_name, "第 2 行示例 user_name 不应为空"
        assert row3_user_name, "第 3 行示例 user_name 不应为空"

    async def test_data_sheet_has_data_validations(self, client, admin_token):
        """数据 sheet 的部门列和角色列包含 DataValidation。"""
        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        ws = wb["数据"]
        # 至少有 2 个 DataValidation（dept + role）
        assert len(ws.data_validations.dataValidation) >= 2, (
            "「数据」sheet 应至少有 2 个 DataValidation（dept + role）"
        )

        # 校验 formula1 引用了字典 sheet
        formulas = []
        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                formulas.append(str(dv.formula1))
        joined = "|".join(formulas)
        assert "部门字典" in joined, (
            f"DataValidation 未引用「部门字典」sheet: {formulas}"
        )
        assert "角色字典" in joined, (
            f"DataValidation 未引用「角色字典」sheet: {formulas}"
        )


class TestTemplateDictSheets:
    """部门和角色字典 sheet 从数据库实时生成。"""

    async def test_dept_dict_sheet_has_header_and_timestamp(
        self, client, admin_token, db_session
    ):
        """「部门字典」sheet：row 1 生成时间 + row 2 表头 + row 3+ 数据行。

        实时查询语义由 ``test_user_template_service.py`` 直接测 service 验证
        （HTTP 测试的 ASGITransport 与 db_session fixture 不共享事务，seed 数据
        对 endpoint 不可见 — 此处只验证 sheet 结构）。
        """
        _dept, _role = await _seed_dept_and_role(db_session)
        await db_session.commit()

        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        ws = wb["部门字典"]

        # 第一行记录模板生成时间。
        row1 = [cell.value for cell in ws[1]]
        assert any("生成时间" in str(c) for c in row1 if c), (
            f"「部门字典」sheet row 1 应是生成时间标注: {row1}"
        )

        # 表头在 row 2：dept_name + full_path + dept_id + status
        header = [cell.value for cell in ws[2]]
        assert "dept_name" in header, f"部门字典 sheet 缺 dept_name 列: {header}"
        assert "full_path" in header, f"部门字典 sheet 缺 full_path 列: {header}"

    async def test_role_dict_sheet_has_header_and_timestamp(
        self, client, admin_token, db_session
    ):
        """「角色字典」sheet：row 1 生成时间 + row 2 表头 + row 3+ 数据行。"""
        _dept, _role = await _seed_dept_and_role(db_session)
        await db_session.commit()

        response = await client.get(
            "/system/user/import/template",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        wb = _load_xlsx(response.content)
        ws = wb["角色字典"]

        # row 1 是生成时间标注
        row1 = [cell.value for cell in ws[1]]
        assert any("生成时间" in str(c) for c in row1 if c), (
            f"「角色字典」sheet row 1 应是生成时间标注: {row1}"
        )

        # 表头在 row 2
        header = [cell.value for cell in ws[2]]
        assert "role_code" in header, f"角色字典 sheet 缺 role_code 列: {header}"
        assert "role_name" in header, f"角色字典 sheet 缺 role_name 列: {header}"
