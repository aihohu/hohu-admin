"""用户导入模板 service 行为测试。

直接调 ``generate_import_template(db_session)`` 验证 service 层契约：
- 部门字典 sheet 实时查 sys_dept（seed 后能立即出现在 xlsx）
- 角色字典 sheet 实时查 sys_role
- full_path 按 ancestors 链正确拼接
- 禁用 dept / role 不进字典（仅 status='1'）

HTTP 契约测试在 ``test_user_import_template_api.py``；本文件聚焦 service 业务逻辑。
"""

import re
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.modules.system.models.dept import Dept
from app.modules.system.models.role import Role
from app.modules.system.user.template_service import generate_import_template

#: 四个 sheet 名称。
_EXPECTED_SHEET_NAMES: tuple[str, ...] = (
    "数据",
    "说明",
    "部门字典",
    "角色字典",
)

#: 数据 sheet 列顺序
_EXPECTED_DATA_COLUMNS: tuple[str, ...] = (
    "user_name",
    "nickname",
    "user_email",
    "user_phone",
    "dept_input",
    "role_input",
    "user_gender",
    "status",
)


def _load_xlsx(content: bytes):
    """加载 xlsx bytes → Workbook。"""
    return load_workbook(BytesIO(content), data_only=False)


@pytest.fixture
async def seed_dept_tree(db_session) -> tuple[Dept, Dept, Dept]:
    """建 3 级 dept 树验证 full_path 拼接。

    总公司 (ancestors='0')
      └─ 研发中心 (ancestors='0,{总公司.id}')
           └─ 前端部 (ancestors='0,{总公司.id},{研发中心.id}')
    """
    root = Dept(
        dept_name="模板总公司",
        parent_id=None,
        ancestors="0",
        status="1",
        order_num=1,
    )
    db_session.add(root)
    await db_session.flush()

    mid = Dept(
        dept_name="模板研发中心",
        parent_id=root.dept_id,
        ancestors=f"0,{root.dept_id}",
        status="1",
        order_num=1,
    )
    db_session.add(mid)
    await db_session.flush()

    leaf = Dept(
        dept_name="模板前端部",
        parent_id=mid.dept_id,
        ancestors=f"0,{root.dept_id},{mid.dept_id}",
        status="1",
        order_num=1,
    )
    db_session.add(leaf)
    await db_session.flush()
    return root, mid, leaf


@pytest.fixture
async def seed_disabled_dept_and_role(db_session) -> tuple[Dept, Role]:
    """建 1 个禁用 dept + 1 个禁用 role，验证 status='2' 不进字典。"""
    disabled_dept = Dept(
        dept_name="模板禁用部门",
        parent_id=None,
        ancestors="0",
        status="2",  # 禁用
        order_num=999,
    )
    db_session.add(disabled_dept)
    disabled_role = Role(
        role_name="模板禁用角色",
        role_code="R_TEMPLATE_DISABLED",
        status="2",  # 禁用
        data_scope="1",
    )
    db_session.add(disabled_role)
    await db_session.flush()
    return disabled_dept, disabled_role


@pytest.fixture
async def seed_role(db_session) -> Role:
    """建 1 个启用 role 验证出现在字典 sheet。"""
    role = Role(
        role_name="模板启用角色",
        role_code="R_TEMPLATE_ENABLED",
        status="1",
        data_scope="1",
    )
    db_session.add(role)
    await db_session.flush()
    return role


class TestTemplateSheetStructure:
    """4 sheet 存在 + 顺序 + 数据 sheet 列 / 示例 / DataValidation。"""

    async def test_has_four_sheets_in_expected_order(self, db_session):
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        assert wb.sheetnames == list(_EXPECTED_SHEET_NAMES)

    async def test_data_sheet_columns(self, db_session):
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["数据"]
        header = [cell.value for cell in ws[1]]
        assert header == list(_EXPECTED_DATA_COLUMNS)

    async def test_data_sheet_has_two_example_rows(self, db_session):
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["数据"]
        # 第 2、3 行示例（user_name 列非空）
        assert ws.cell(row=2, column=1).value
        assert ws.cell(row=3, column=1).value

    async def test_data_sheet_has_data_validations_referencing_dict_sheets(
        self, db_session
    ):
        """部门列和角色列的 DataValidation 引用对应字典 sheet。"""
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["数据"]
        assert len(ws.data_validations.dataValidation) >= 2

        formulas = [
            str(dv.formula1) for dv in ws.data_validations.dataValidation if dv.formula1
        ]
        joined = "|".join(formulas)
        assert "部门字典" in joined, (
            f"DataValidation 未引用「部门字典」sheet: {formulas}"
        )
        assert "角色字典" in joined, (
            f"DataValidation 未引用「角色字典」sheet: {formulas}"
        )


class TestDeptDictSheetRealtime:
    """部门字典 sheet 从 sys_dept 实时生成。"""

    async def test_seed_dept_appears_in_dict_sheet(self, db_session, seed_dept_tree):
        """seed 的 3 级 dept 树应全部出现在「部门字典」sheet。"""
        root, mid, leaf = seed_dept_tree
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["部门字典"]

        # row 1 生成时间 / row 2 表头 / row 3+ 数据
        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        dept_names = {row[0] for row in all_rows if row and row[0]}
        assert root.dept_name in dept_names
        assert mid.dept_name in dept_names
        assert leaf.dept_name in dept_names

    async def test_full_path_built_from_ancestors_chain(
        self, db_session, seed_dept_tree
    ):
        """full_path 由祖先链和当前部门名称拼接。"""
        root, mid, leaf = seed_dept_tree
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["部门字典"]

        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        full_path_by_name: dict[str, str] = {
            row[0]: row[1] for row in all_rows if row and row[0] and row[1]
        }

        # 根节点：full_path == dept_name（无祖先）
        assert full_path_by_name[root.dept_name] == root.dept_name
        # 中间节点：root_name/mid_name
        assert full_path_by_name[mid.dept_name] == f"{root.dept_name}/{mid.dept_name}"
        # 叶子节点：root_name/mid_name/leaf_name
        assert (
            full_path_by_name[leaf.dept_name]
            == f"{root.dept_name}/{mid.dept_name}/{leaf.dept_name}"
        )

    async def test_disabled_dept_excluded_from_dict_sheet(
        self, db_session, seed_disabled_dept_and_role
    ):
        """status='2' 禁用 dept 不进字典 sheet。"""
        disabled_dept, _role = seed_disabled_dept_and_role
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["部门字典"]

        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        dept_names = {row[0] for row in all_rows if row and row[0]}
        assert disabled_dept.dept_name not in dept_names, (
            f"禁用 dept 不应出现在字典 sheet: {disabled_dept.dept_name}"
        )

    async def test_dept_dict_sheet_has_timestamp_at_row_1(self, db_session):
        """第一行记录生成时间。"""
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["部门字典"]
        row1 = [cell.value for cell in ws[1]]
        assert any("生成时间" in str(c) for c in row1 if c), (
            f"row 1 应是生成时间标注: {row1}"
        )
        # 校验时间格式 YYYY-MM-DD HH:MM
        all_text = " ".join(str(c) for c in row1 if c)
        assert re.search(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}", all_text), (
            f"生成时间格式错: {all_text}"
        )


class TestRoleDictSheetRealtime:
    """角色字典 sheet 从 sys_role 实时生成。"""

    async def test_seed_role_appears_in_dict_sheet(self, db_session, seed_role):
        """seed 的 role 应出现在字典 sheet。"""
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["角色字典"]

        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        role_codes = {row[0] for row in all_rows if row and row[0]}
        assert seed_role.role_code in role_codes

    async def test_disabled_role_excluded_from_dict_sheet(
        self, db_session, seed_disabled_dept_and_role
    ):
        """status='2' 禁用 role 不进字典 sheet。"""
        _dept, disabled_role = seed_disabled_dept_and_role
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["角色字典"]

        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        role_codes = {row[0] for row in all_rows if row and row[0]}
        assert disabled_role.role_code not in role_codes, (
            f"禁用 role 不应出现在字典 sheet: {disabled_role.role_code}"
        )

    async def test_role_dict_sheet_has_timestamp_at_row_1(self, db_session):
        """第一行记录生成时间。"""
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["角色字典"]
        row1 = [cell.value for cell in ws[1]]
        assert any("生成时间" in str(c) for c in row1 if c), (
            f"row 1 应是生成时间标注: {row1}"
        )


class TestInstructionSheet:
    """说明 sheet 包含字段说明、必填、取值范围和冲突处理。"""

    async def test_instruction_sheet_has_field_descriptions(self, db_session):
        """说明 sheet 至少覆盖 8 个字段 + 表头行。"""
        xlsx_bytes = await generate_import_template(db_session)
        wb = _load_xlsx(xlsx_bytes)
        ws = wb["说明"]
        all_rows = list(ws.iter_rows(values_only=True))
        # 第 1 行表头 + 至少 8 个字段说明行
        assert len(all_rows) >= 9, f"说明 sheet 行数不足: {len(all_rows)}"
        # 表头含「字段名」
        header = all_rows[0]
        assert "字段名" in [str(c) for c in header if c], f"说明 sheet 表头错: {header}"
        # 必填字段 user_name 在说明里
        all_text = " ".join(str(c) for row in all_rows for c in row if c)
        assert "user_name" in all_text
        assert "dept_input" in all_text
